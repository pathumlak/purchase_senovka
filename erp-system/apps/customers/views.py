import json
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Case, When, F, Value, DecimalField
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from .models import Customer, CustomerLedger, CustomerProductPrice, add_ledger_entry
from .forms import CustomerForm, CustomerProductPriceForm
from apps.production.models import Product
from apps.billing.models import Payment
from apps.billing.views import _parse_cheques
from django.db import transaction
from erp.utils import log_activity, resolve_date_range


def _is_superadmin(request):
    try:
        return request.user.profile.is_superadmin()
    except Exception:
        return False


# ── Customer CRUD ──────────────────────────────────────────────

def customer_list(request):
    # Balance totals always reflect every customer, regardless of the
    # "added this month" filter below -- these are running-balance figures,
    # not something tied to when the customer record was created.
    all_customers_qs = Customer.objects.all()
    totals = all_customers_qs.aggregate(
        total_credit=Sum(
            Case(When(balance__gt=0, then=F('balance')),
                 default=Value(0),
                 output_field=DecimalField(max_digits=12, decimal_places=2))
        ),
        total_outstanding=Sum(
            Case(When(balance__lt=0, then=F('balance')),
                 default=Value(0),
                 output_field=DecimalField(max_digits=12, decimal_places=2))
        ),
    )

    date_from, date_to = resolve_date_range(request)
    customers_qs = Customer.objects.order_by('-created_at')
    if date_from:
        customers_qs = customers_qs.filter(created_at__date__gte=date_from)
    if date_to:
        customers_qs = customers_qs.filter(created_at__date__lte=date_to)

    paginator = Paginator(customers_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'customers/customer_list.html', {
        'customers': page_obj,
        'page_obj': page_obj,
        'total_customers': all_customers_qs.count(),
        'filtered_count': customers_qs.count(),
        'date_from': date_from,
        'date_to': date_to,
        'total_credit': totals['total_credit'] or Decimal('0'),
        'total_outstanding': abs(totals['total_outstanding'] or Decimal('0')),
        'is_superadmin': _is_superadmin(request),
    })


def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            if _is_superadmin(request):
                raw = request.POST.get('max_outstanding_limit', '').strip()
                if raw:
                    try:
                        customer.max_outstanding_limit = Decimal(raw)
                        customer.save(update_fields=['max_outstanding_limit'])
                    except Exception:
                        pass
            log_activity(request, 'customers', 'customer_created',
                         f"Customer created: {customer.name}", reverse('customer_list'),
                         related_id=customer.pk)
            messages.success(request, 'Customer created successfully.')
    return redirect('customer_list')


def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            if _is_superadmin(request):
                raw = request.POST.get('max_outstanding_limit', '').strip()
                if raw:
                    try:
                        customer.max_outstanding_limit = Decimal(raw)
                    except Exception:
                        customer.max_outstanding_limit = None
                else:
                    customer.max_outstanding_limit = None
                customer.save(update_fields=['max_outstanding_limit'])
            log_activity(request, 'customers', 'customer_updated',
                         f"Customer updated: {customer.name}", reverse('customer_list'))
            messages.success(request, 'Customer updated successfully.')
    return redirect('customer_list')


def customer_update_balance(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        try:
            old_balance = customer.balance
            new_balance = Decimal(request.POST.get('balance', '0'))
            customer.balance = new_balance
            customer.save(update_fields=['balance', 'updated_at'])
            delta = new_balance - old_balance
            if delta != 0:
                add_ledger_entry(
                    customer,
                    date=timezone.localdate(),
                    description='Manual Balance Adjustment',
                    transaction_type=CustomerLedger.MANUAL_ADJUSTMENT,
                    debit=max(Decimal('0'), -delta),
                    credit=max(Decimal('0'), delta),
                )
            log_activity(request, 'customers', 'balance_updated',
                         f"Balance updated for {customer.name}: {customer.balance}",
                         reverse('customer_list'))
            messages.success(request, f'Balance updated for {customer.name}.')
        except Exception:
            messages.error(request, 'Invalid balance value.')
    return redirect('customer_list')


_SETTLE_METHODS = [
    ('FULL_CASH',      'Full Cash'),
    ('FULL_CHEQUE',    'Full Cheque'),
    ('PARTIAL_CASH',   'Partial Cash'),
    ('PARTIAL_CHEQUE', 'Partial Cheque'),
    ('MIXED',          'Mixed (Cash + Cheque)'),
]


def customer_settle_balance(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if customer.balance >= 0:
        messages.info(request, f'{customer.name} has no outstanding balance to settle.')
        return redirect('customer_list')

    outstanding = abs(customer.balance)

    if request.method == 'GET':
        return render(request, 'customers/settle_balance.html', {
            'customer': customer,
            'outstanding': outstanding,
            'today': timezone.localdate().isoformat(),
            'method_choices': _SETTLE_METHODS,
        })

    # POST — process settlement
    try:
        with transaction.atomic():
            customer_obj = Customer.objects.select_for_update().get(pk=pk)
            if customer_obj.balance >= 0:
                messages.error(request, 'No outstanding balance.')
                return redirect('customer_list')

            outstanding = abs(customer_obj.balance)
            method = request.POST.get('settlement_method', '')
            valid = {'FULL_CASH', 'FULL_CHEQUE', 'PARTIAL_CASH', 'PARTIAL_CHEQUE', 'MIXED'}
            if method not in valid:
                raise ValueError('Select a valid settlement method.')

            def _d(val):
                try:
                    return Decimal(str(val or '0')).quantize(Decimal('0.01'))
                except (InvalidOperation, TypeError):
                    return Decimal('0')

            cash_in        = Decimal('0')
            cheque_in      = Decimal('0')
            cheque_records = []
            if method in ('FULL_CASH', 'PARTIAL_CASH', 'MIXED'):
                cash_in = _d(request.POST.get('cash_amount'))
            if method in ('FULL_CHEQUE', 'PARTIAL_CHEQUE', 'MIXED'):
                cheque_records, cheque_in = _parse_cheques(request.POST, customer_obj.name)

            collected = cash_in + cheque_in
            if collected <= 0:
                raise ValueError('Enter a payment amount.')

            if method in ('FULL_CASH', 'FULL_CHEQUE', 'MIXED') and collected < outstanding:
                raise ValueError(
                    f'Full settlement requires Rs. {outstanding:,.2f}. You entered Rs. {collected:,.2f}.'
                )

            applied   = min(collected, outstanding)
            settle_date = timezone.localdate()

            # Create Payment records
            if cash_in > 0:
                senovka = request.POST.get('cash_is_senovka_transfer') == '1'
                Payment.objects.create(
                    customer=customer_obj,
                    payment_date=settle_date,
                    method=Payment.CASH,
                    amount=cash_in,
                    is_senovka_transfer=senovka,
                )

            for record in cheque_records:
                Payment.objects.create(
                    customer=customer_obj,
                    payment_date=settle_date,
                    **record,
                )

            # Update customer balance
            customer_obj.balance += applied
            customer_obj.save(update_fields=['balance', 'updated_at'])

            # Ledger entries
            if cash_in > 0:
                cash_applied = min(cash_in, outstanding)
                add_ledger_entry(
                    customer_obj,
                    date=settle_date,
                    description='Cash Payment (Direct Balance Settlement)',
                    transaction_type=CustomerLedger.PAYMENT_CASH,
                    credit=cash_applied,
                )
            if cheque_in > 0:
                chq_applied = min(cheque_in, max(Decimal('0'), outstanding - min(cash_in, outstanding)))
                add_ledger_entry(
                    customer_obj,
                    date=settle_date,
                    description=f'Cheque Payment (Direct Balance Settlement)',
                    transaction_type=CustomerLedger.PAYMENT_CHEQUE,
                    credit=chq_applied,
                )

            log_activity(request, 'customers', 'balance_settled',
                         f"Balance settled for {customer_obj.name}: Rs. {applied}",
                         reverse('customer_list'))
            messages.success(request, f'Rs. {applied:,.2f} settled for {customer_obj.name}.')
            return redirect('customer_list')

    except ValueError as exc:
        messages.error(request, str(exc))
        return render(request, 'customers/settle_balance.html', {
            'customer': customer,
            'outstanding': outstanding,
            'today': timezone.localdate().isoformat(),
            'method_choices': _SETTLE_METHODS,
        })


def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        try:
            name = customer.name
            customer.delete()
            log_activity(request, 'customers', 'customer_deleted',
                         f"Customer deleted: {name}", reverse('customer_list'))
            messages.success(request, 'Customer deleted successfully.')
        except ProtectedError:
            messages.error(
                request,
                'Cannot delete this customer because it is used in billing records.'
            )
    return redirect('customer_list')


# ── Customer Product Pricing CRUD ──────────────────────────────

def customer_pricing_list(request):
    q = request.GET.get('q', '').strip()
    prices_qs = CustomerProductPrice.objects.select_related(
        'customer', 'product', 'product__category'
    ).order_by('customer__name', 'product__name')
    if q:
        prices_qs = prices_qs.filter(
            Q(customer__name__icontains=q) | Q(product__name__icontains=q)
        )
    total_prices = CustomerProductPrice.objects.count()
    paginator = Paginator(prices_qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    customers = Customer.objects.all().order_by('name')
    products = Product.objects.select_related('category').all().order_by('name')
    return render(request, 'customers/pricing_list.html', {
        'prices': page_obj,
        'page_obj': page_obj,
        'total_prices': total_prices,
        'filtered_count': prices_qs.count(),
        'customers': customers,
        'products': products,
        'q': q,
    })


def pricing_create(request):
    if request.method == 'POST':
        form = CustomerProductPriceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer price created successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    return redirect('customer_pricing_list')


def pricing_update(request, pk):
    price = get_object_or_404(CustomerProductPrice, pk=pk)
    if request.method == 'POST':
        form = CustomerProductPriceForm(request.POST, instance=price)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer price updated successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    return redirect('customer_pricing_list')


def pricing_delete(request, pk):
    price = get_object_or_404(CustomerProductPrice, pk=pk)
    if request.method == 'POST':
        price.delete()
        messages.success(request, 'Customer price deleted successfully.')
    return redirect('customer_pricing_list')


# ── Pricing exports ────────────────────────────────────────────

def _filtered_pricing_qs(request):
    q = request.GET.get('q', '').strip()
    qs = CustomerProductPrice.objects.select_related(
        'customer', 'product', 'product__category'
    ).order_by('customer__name', 'product__name')
    if q:
        qs = qs.filter(
            Q(customer__name__icontains=q) | Q(product__name__icontains=q)
        )
    return qs, q


def pricing_export_excel(request):
    qs, q = _filtered_pricing_qs(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Customer Pricing'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1D4ED8')

    generated_at = timezone.localtime()
    ws.append(['Customer Pricing Report'])
    ws.append(['Generated At', generated_at.strftime('%d %b %Y %I:%M %p')])
    if q:
        ws.append(['Filter', q])
    ws.append(['Total Rows', qs.count()])
    ws.append([])

    headers = ['Customer', 'Product', 'Category', 'Unit Price (Rs.)']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for price in qs:
        ws.append([
            price.customer.name,
            price.product.name,
            price.product.category.name,
            float(price.unit_price),
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or '')) for cell in col
        ) + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = generated_at.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="customer_pricing_{ts}.xlsx"'
    wb.save(response)
    return response


def pricing_export_pdf(request):
    qs, q = _filtered_pricing_qs(request)
    from django.template.loader import get_template
    context = {
        'prices': list(qs),
        'q': q,
        'generated_at': timezone.localtime(),
        'total': qs.count(),
    }
    template = get_template('customers/pricing_pdf.html')
    html = template.render(context)
    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="customer_pricing_{ts}.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('PDF generation error', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'customers/pricing_pdf.html', context)


# ── API: live customer search ──────────────────────────────────

def customer_search_api(request):
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'customers': []})

    qs = Customer.objects.filter(
        Q(name__icontains=q) | Q(address__icontains=q)
    ).order_by('name')[:30]

    data = [
        {
            'id': c.id,
            'name': c.name,
            'address': c.address or '',
            'balance': str(c.balance),
            'max_outstanding_limit': str(c.max_outstanding_limit) if c.max_outstanding_limit is not None else '',
            'created_at': c.created_at.strftime('%b %d, %Y'),
            'updated_at': c.updated_at.strftime('%b %d, %Y %H:%M'),
        }
        for c in qs
    ]
    return JsonResponse({'customers': data, 'count': len(data)})


# ── API endpoint for getting customer price ────────────────────

def get_customer_price(request):
    """
    AJAX endpoint: GET /customers/api/price/?customer_id=1&product_id=2
    Returns the price for a given customer-product pair.
    """
    customer_id = request.GET.get('customer_id')
    product_id = request.GET.get('product_id')
    if not customer_id or not product_id:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    try:
        customer = Customer.objects.get(pk=customer_id)
        product = Product.objects.get(pk=product_id)
    except (Customer.DoesNotExist, Product.DoesNotExist):
        return JsonResponse({'error': 'Not found'}, status=404)

    price = CustomerProductPrice.get_price_for_customer(customer, product)
    return JsonResponse({
        'customer': customer.name,
        'product': product.name,
        'unit_price': str(price),
        'is_custom': CustomerProductPrice.objects.filter(
            customer=customer, product=product
        ).exists(),
    })


# ── Bulk Pricing ───────────────────────────────────────────────

def bulk_pricing_view(request):
    customers = Customer.objects.order_by('name')
    return render(request, 'customers/pricing_bulk.html', {'customers': customers})


def customer_products_api(request):
    """
    GET /customers/api/customer-products/?customer_id=X
    Returns all products with existing custom prices for the given customer.
    """
    customer_id = request.GET.get('customer_id')
    if not customer_id:
        return JsonResponse({'error': 'customer_id required'}, status=400)
    try:
        customer = Customer.objects.get(pk=customer_id)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)

    existing = {
        cp.product_id: str(cp.unit_price)
        for cp in CustomerProductPrice.objects.filter(customer=customer)
    }

    products = (
        Product.objects
        .select_related('category')
        .order_by('category__name', 'name')
    )

    data = [
        {
            'id': p.id,
            'name': p.name,
            'category': p.category.name,
            'unit_price': existing.get(p.id, ''),
        }
        for p in products
    ]
    return JsonResponse({'customer': customer.name, 'products': data})


@require_POST
def bulk_pricing_save(request):
    """
    POST /customers/pricing/bulk/save/
    Body JSON: { "customer_id": 1, "prices": [{"product_id": 1, "unit_price": "100.00"}, ...] }
    Empty unit_price string = delete existing custom price for that product.
    """
    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    customer_id = payload.get('customer_id')
    prices = payload.get('prices', [])

    if not customer_id:
        return JsonResponse({'error': 'customer_id required'}, status=400)
    try:
        customer = Customer.objects.get(pk=customer_id)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)

    saved = deleted = skipped = 0
    errors = []

    for item in prices:
        product_id = item.get('product_id')
        raw_price = str(item.get('unit_price', '')).strip()

        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            errors.append(f'Product {product_id} not found')
            continue

        if raw_price == '':
            count, _ = CustomerProductPrice.objects.filter(
                customer=customer, product=product
            ).delete()
            deleted += count
        else:
            try:
                unit_price = Decimal(raw_price)
                if unit_price < 0:
                    raise ValueError('negative')
            except (InvalidOperation, ValueError):
                errors.append(f'Invalid price for product {product.name}: {raw_price}')
                skipped += 1
                continue

            CustomerProductPrice.objects.update_or_create(
                customer=customer,
                product=product,
                defaults={'unit_price': unit_price},
            )
            saved += 1

    return JsonResponse({
        'ok': True,
        'saved': saved,
        'deleted': deleted,
        'skipped': skipped,
        'errors': errors,
        'message': f'{saved} price(s) saved, {deleted} removed.',
    })


# ── Customer Ledger ────────────────────────────────────────────

def _ledger_filtered_qs(customer, request):
    qs = CustomerLedger.objects.filter(customer=customer).select_related('bill').order_by('date', 'created_at')
    date_from, date_to = resolve_date_range(request)
    ttype = request.GET.get('type', '').strip()
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if ttype:
        qs = qs.filter(transaction_type=ttype)
    return qs, date_from, date_to, ttype


def customer_ledger(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    entries_qs, date_from, date_to, ttype = _ledger_filtered_qs(customer, request)

    totals = entries_qs.aggregate(total_debit=Sum('debit'), total_credit=Sum('credit'))
    total_debit = totals['total_debit'] or Decimal('0')
    total_credit = totals['total_credit'] or Decimal('0')

    paginator = Paginator(entries_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'customers/customer_ledger.html', {
        'customer': customer,
        'page_obj': page_obj,
        'entries': page_obj,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'date_from': date_from,
        'date_to': date_to,
        'ttype': ttype,
        'transaction_types': CustomerLedger.TRANSACTION_TYPES,
    })


def customer_ledger_export_excel(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    entries_qs, date_from, date_to, ttype = _ledger_filtered_qs(customer, request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    totals = entries_qs.aggregate(total_debit=Sum('debit'), total_credit=Sum('credit'))
    total_debit = totals['total_debit'] or Decimal('0')
    total_credit = totals['total_credit'] or Decimal('0')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ledger'
    gen = timezone.localtime()

    # ── title block ──
    ws.append([f'Customer Ledger — {customer.name}'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Current Balance', f'Rs. {customer.balance:.2f}'])
    if date_from or date_to:
        ws.append(['Period', f'{date_from or "—"} to {date_to or "—"}'])
    ws.append([])

    # ── summary ──
    ws.append(['Total Debit (DR)', '', f'Rs. {total_debit:.2f}'])
    ws.append(['Total Credit (CR)', '', f'Rs. {total_credit:.2f}'])
    ws.append([])

    # ── headers ──
    headers = ['#', 'Date', 'Bill No.', 'Description', 'Debit (DR)', 'Credit (CR)', 'Balance']
    ws.append(headers)
    hrow = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    # ── rows ──
    for i, e in enumerate(entries_qs, 1):
        ws.append([
            i,
            e.date.strftime('%d-%m-%Y'),
            e.bill_number or '—',
            e.description,
            float(e.debit) if e.debit else '',
            float(e.credit) if e.credit else '',
            float(e.balance),
        ])

    # ── totals row ──
    ws.append([])
    tfont = Font(bold=True)
    ws.append(['', '', '', 'TOTAL', float(total_debit), float(total_credit), float(customer.balance)])
    for ci in range(1, 8):
        ws.cell(row=ws.max_row, column=ci).font = tfont

    # column widths
    for col, w in zip('ABCDEFG', [4, 13, 20, 45, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = customer.name.replace(' ', '_')[:30]
    ts = gen.strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="ledger_{safe_name}_{ts}.xlsx"'
    wb.save(response)
    return response


def customer_ledger_export_pdf(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    entries_qs, date_from, date_to, ttype = _ledger_filtered_qs(customer, request)

    totals = entries_qs.aggregate(total_debit=Sum('debit'), total_credit=Sum('credit'))
    total_debit = totals['total_debit'] or Decimal('0')
    total_credit = totals['total_credit'] or Decimal('0')

    from django.template.loader import get_template
    context = {
        'customer': customer,
        'entries': list(entries_qs),
        'total_debit': total_debit,
        'total_credit': total_credit,
        'date_from': date_from,
        'date_to': date_to,
        'ttype': ttype,
        'generated_at': timezone.localtime(),
    }
    template = get_template('customers/ledger_pdf.html')
    html = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        safe_name = customer.name.replace(' ', '_')[:30]
        ts = timezone.localtime().strftime('%Y%m%d')
        response['Content-Disposition'] = f'attachment; filename="ledger_{safe_name}_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'customers/ledger_pdf.html', context)


# ── Customer Balances Report ───────────────────────────────────

def _customer_balances_qs(request):
    """All customers + their current balance, filterable by name search and
    balance type (owing = negative, credit = positive, settled = zero)."""
    q = (request.GET.get('q') or '').strip()
    balance_type = (request.GET.get('balance_type') or '').strip()

    qs = Customer.objects.order_by('name')
    if q:
        qs = qs.filter(name__icontains=q)
    if balance_type == 'owing':
        qs = qs.filter(balance__lt=0)
    elif balance_type == 'credit':
        qs = qs.filter(balance__gt=0)
    elif balance_type == 'settled':
        qs = qs.filter(balance=0)

    return qs, q, balance_type


def customer_balances_report(request):
    customers_qs, q, balance_type = _customer_balances_qs(request)

    totals = customers_qs.aggregate(
        total_credit=Sum(
            Case(When(balance__gt=0, then=F('balance')),
                 default=Value(0),
                 output_field=DecimalField(max_digits=14, decimal_places=2))
        ),
        total_outstanding=Sum(
            Case(When(balance__lt=0, then=F('balance')),
                 default=Value(0),
                 output_field=DecimalField(max_digits=14, decimal_places=2))
        ),
    )
    total_credit = totals['total_credit'] or Decimal('0')
    total_outstanding = abs(totals['total_outstanding'] or Decimal('0'))

    paginator = Paginator(customers_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'customers/balances_report.html', {
        'customers':         page_obj,
        'page_obj':          page_obj,
        'total_customers':   customers_qs.count(),
        'total_credit':      total_credit,
        'total_outstanding': total_outstanding,
        'q':                 q,
        'balance_type':      balance_type,
        'today':             timezone.localdate(),
        'query_string':      query_params.urlencode(),
    })


def customer_balances_export_excel(request):
    customers_qs, q, balance_type = _customer_balances_qs(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    total_credit = Decimal('0')
    total_outstanding = Decimal('0')

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Customer Balances'
    gen = timezone.localtime()

    ws.append(['Customer Balances Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['As Of', gen.strftime('%d %b %Y')])
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Customers', customers_qs.count()])
    ws.append([])

    headers = ['#', 'Customer', 'Balance (Rs.)', 'Status']
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, c in enumerate(customers_qs, 1):
        if c.balance > 0:
            status = 'Credit (we owe)'
            total_credit += c.balance
        elif c.balance < 0:
            status = 'Outstanding (owes us)'
            total_outstanding += abs(c.balance)
        else:
            status = 'Settled'
        ws.append([i, c.name, float(c.balance), status])

    ws.append([])
    tfont = Font(bold=True)
    ws.append(['', 'TOTAL CREDIT (we owe)', float(total_credit), ''])
    for ci in range(1, 5):
        ws.cell(row=ws.max_row, column=ci).font = tfont
    ws.append(['', 'TOTAL OUTSTANDING (owed to us)', float(total_outstanding), ''])
    for ci in range(1, 5):
        ws.cell(row=ws.max_row, column=ci).font = tfont

    for col, w in zip('ABCD', [4, 32, 16, 22]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="customer_balances_{ts}.xlsx"'
    wb.save(response)
    return response


def customer_balances_export_pdf(request):
    customers_qs, q, balance_type = _customer_balances_qs(request)

    totals = customers_qs.aggregate(
        total_credit=Sum(
            Case(When(balance__gt=0, then=F('balance')),
                 default=Value(0),
                 output_field=DecimalField(max_digits=14, decimal_places=2))
        ),
        total_outstanding=Sum(
            Case(When(balance__lt=0, then=F('balance')),
                 default=Value(0),
                 output_field=DecimalField(max_digits=14, decimal_places=2))
        ),
    )

    from django.template.loader import get_template
    context = {
        'customers':         list(customers_qs),
        'total_customers':   customers_qs.count(),
        'total_credit':      totals['total_credit'] or Decimal('0'),
        'total_outstanding': abs(totals['total_outstanding'] or Decimal('0')),
        'today':             timezone.localdate(),
        'generated_at':      timezone.localtime(),
    }
    template = get_template('customers/balances_report_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="customer_balances_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'customers/balances_report_pdf.html', context)
