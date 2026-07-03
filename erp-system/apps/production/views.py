from django.shortcuts import render, get_object_or_404, redirect


def _fmt_qty(q):
    return f'{q:.2f}'.rstrip('0').rstrip('.')
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.http import JsonResponse, HttpResponse
from django.template.loader import get_template
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.db import transaction
from django.db.models import Sum
from .models import (
    ProductCategory, Product, DailyRunningMachine, Machine, DailyWorkAssignment,
    ProductionEntry, Employee, ShiftProductionLog, ProductionItem, ShiftTemplate, TargetLog,
)
from .forms import (
    ProductCategoryForm, ProductForm, DailyRunningMachineForm, MachineForm,
    DailyWorkAssignmentForm, EmployeeForm, ProductionItemForm, ShiftTemplateForm, TargetLogForm,
)
from apps.customers.models import Customer, CustomerProductPrice
from apps.billing.models import Bill, BillItem
from erp.utils import log_activity, resolve_date_range, current_month_bounds


# ── Category CRUD ──────────────────────────────────────────────

def _category_filters_from_request(request):
    search = (request.GET.get('q') or '').strip()
    stock_mode = request.GET.get('stock_mode', 'all')
    if stock_mode not in {'all', 'with_products', 'without_products'}:
        stock_mode = 'all'
    date_from, date_to = resolve_date_range(request)
    return search, stock_mode, date_from, date_to


def _category_base_queryset():
    return ProductCategory.objects.annotate(product_count=Count('products')).all()


def _apply_category_filters(queryset, search, stock_mode, date_from='', date_to=''):
    if search:
        queryset = queryset.filter(name__icontains=search)

    if stock_mode == 'with_products':
        queryset = queryset.filter(product_count__gt=0)
    elif stock_mode == 'without_products':
        queryset = queryset.filter(product_count=0)

    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    return queryset.order_by('-created_at')


def _filtered_category_queryset(request):
    search, stock_mode, date_from, date_to = _category_filters_from_request(request)
    queryset = _apply_category_filters(_category_base_queryset(), search, stock_mode, date_from, date_to)
    return queryset, search, stock_mode, date_from, date_to

def category_list(request):
    categories_qs, search, stock_mode, date_from, date_to = _filtered_category_queryset(request)
    paginator = Paginator(categories_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    total_products = Product.objects.count()
    total_categories = ProductCategory.objects.count()
    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'production/categories/category_list.html', {
        'categories': page_obj,
        'page_obj': page_obj,
        'total_products': total_products,
        'total_categories': total_categories,
        'search_query': search,
        'stock_mode': stock_mode,
        'date_from': date_from,
        'date_to': date_to,
        'filtered_count': categories_qs.count(),
        'query_string': query_params.urlencode(),
    })


def category_export_pdf(request):
    categories, search, stock_mode, date_from, date_to = _filtered_category_queryset(request)

    stock_mode_label = {
        'all': 'All',
        'with_products': 'With Products',
        'without_products': 'Without Products',
    }[stock_mode]

    context = {
        'categories': categories,
        'search_query': search,
        'stock_mode_label': stock_mode_label,
        'generated_at': timezone.localtime(),
    }

    template = get_template('production/categories/category_report_pdf.html')
    html = template.render(context)

    try:
        from xhtml2pdf import pisa

        response = HttpResponse(content_type='application/pdf')
        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="categories_{timestamp}.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('PDF generation error', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/categories/category_report_pdf.html', context)


def category_export_excel(request):
    categories, search, stock_mode, date_from, date_to = _filtered_category_queryset(request)

    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse('Excel export dependency missing: openpyxl', status=500)

    stock_mode_label = {
        'all': 'All',
        'with_products': 'With Products',
        'without_products': 'Without Products',
    }[stock_mode]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Categories'

    sheet.append(['Category Report'])
    sheet.append(['Generated At', timezone.localtime().strftime('%d %b %Y %I:%M %p')])
    sheet.append(['Search', search or 'All'])
    sheet.append(['Stock Mode', stock_mode_label])
    sheet.append(['Total Rows', categories.count()])
    sheet.append([])
    sheet.append(['Category', 'Description', 'Products', 'Created At'])

    for category in categories:
        sheet.append([
            category.name,
            category.description or '',
            category.product_count,
            timezone.localtime(category.created_at).strftime('%d %b %Y %I:%M %p'),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="categories_{timestamp}.xlsx"'
    workbook.save(response)
    return response


def category_create(request):
    if request.method == 'POST':
        form = ProductCategoryForm(request.POST)
        if form.is_valid():
            cat = form.save()
            log_activity(request, 'production', 'category_created',
                         f"Category created: {cat.name}", reverse('category_list'),
                         related_id=cat.pk)
            messages.success(request, 'Category created successfully.')
            return redirect('category_list')
    return redirect('category_list')


def category_update(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    if request.method == 'POST':
        form = ProductCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            log_activity(request, 'production', 'category_updated',
                         f"Category updated: {category.name}", reverse('category_list'))
            messages.success(request, 'Category updated successfully.')
    return redirect('category_list')


def category_delete(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    if request.method == 'POST':
        try:
            name = category.name
            category.delete()
            log_activity(request, 'production', 'category_deleted',
                         f"Category deleted: {name}", reverse('category_list'))
            messages.success(request, 'Category deleted successfully.')
        except ProtectedError:
            messages.error(
                request,
                'Cannot delete this category because related products are already used in billing records.'
            )
    return redirect('category_list')


# ── Product CRUD ───────────────────────────────────────────────

def _save_customer_prices(request, product):
    """Process inline customer pricing rows from the form."""
    customer_ids = request.POST.getlist('customer_id[]')
    customer_prices = request.POST.getlist('customer_price[]')

    # Collect valid customer-price pairs from submit
    submitted_pairs = {}
    for cid, price in zip(customer_ids, customer_prices):
        if cid and price:
            try:
                submitted_pairs[int(cid)] = Decimal(price)
            except (ValueError, InvalidOperation):
                continue

    # Remove old prices not in submission
    CustomerProductPrice.objects.filter(product=product).exclude(
        customer_id__in=submitted_pairs.keys()
    ).delete()

    # Create or update prices
    for customer_id, unit_price in submitted_pairs.items():
        CustomerProductPrice.objects.update_or_create(
            customer_id=customer_id,
            product=product,
            defaults={'unit_price': unit_price},
        )


def product_list(request):
    q = (request.GET.get('q') or '').strip()
    date_from, date_to = resolve_date_range(request)
    products_qs = Product.objects.select_related('category').prefetch_related(
        'customer_prices__customer'
    ).order_by('category__name', 'name')
    if q:
        products_qs = products_qs.filter(
            Q(name__icontains=q) | Q(category__name__icontains=q) | Q(size__icontains=q)
        )
    if date_from:
        products_qs = products_qs.filter(created_at__date__gte=date_from)
    if date_to:
        products_qs = products_qs.filter(created_at__date__lte=date_to)
    total_products = Product.objects.count()
    paginator = Paginator(products_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    categories = ProductCategory.objects.all().order_by('name')
    customers = Customer.objects.all().order_by('name')

    # Build customer pricing data per product for JS
    product_customer_prices = {}
    for product in page_obj:
        product_customer_prices[product.id] = [
            {'customer_id': cp.customer_id, 'customer_name': cp.customer.name, 'unit_price': str(cp.unit_price)}
            for cp in product.customer_prices.all()
        ]

    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'production/products/product_list.html', {
        'products': page_obj,
        'page_obj': page_obj,
        'total_products': total_products,
        'categories': categories,
        'customers': customers,
        'product_customer_prices': product_customer_prices,
        'q': q,
        'date_from': date_from,
        'date_to': date_to,
        'query_string': query_params.urlencode(),
    })


def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            _save_customer_prices(request, product)
            log_activity(request, 'production', 'product_created',
                         f"Product created: {product.name} | Category: {product.category.name}",
                         reverse('product_list'), related_id=product.pk)
            messages.success(request, 'Product created successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    return redirect('product_list')


def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            _save_customer_prices(request, product)
            log_activity(request, 'production', 'product_updated',
                         f"Product updated: {product.name} | Category: {product.category.name}",
                         reverse('product_list'))
            messages.success(request, 'Product updated successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    return redirect('product_list')


def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        try:
            name = product.name
            product.delete()
            log_activity(request, 'production', 'product_deleted',
                         f"Product deleted: {name}", reverse('product_list'))
            messages.success(request, 'Product deleted successfully.')
        except ProtectedError:
            messages.error(
                request,
                'Cannot delete this product because it is used in billing records.'
            )
    return redirect('product_list')


# ── Daily Running Machine CRUD ───────────────────────────────

def _daily_machine_filters_from_request(request):
    month = (request.GET.get('month') or '').strip()
    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()

    # Default to the current month only on a fresh visit (none of month/
    # start_date/end_date present) -- never alongside an explicit `month`,
    # which would otherwise combine into an impossible date range.
    if 'month' not in request.GET and 'start_date' not in request.GET and 'end_date' not in request.GET:
        start_date_raw, end_date_raw = current_month_bounds()

    machine_name = (request.GET.get('machine_name') or '').strip()
    machine_operator = (request.GET.get('machine_operator') or '').strip()
    item_id = (request.GET.get('item_id') or '').strip()
    q = (request.GET.get('q') or '').strip()

    start_date = parse_date(start_date_raw) if start_date_raw else None
    end_date = parse_date(end_date_raw) if end_date_raw else None

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    if item_id and not item_id.isdigit():
        item_id = ''

    if month:
        try:
            year_str, month_str = month.split('-')
            month_year = int(year_str)
            month_number = int(month_str)
            if month_number < 1 or month_number > 12:
                month = ''
                month_year = None
                month_number = None
        except (ValueError, TypeError):
            month = ''
            month_year = None
            month_number = None
    else:
        month_year = None
        month_number = None

    return {
        'month': month,
        'month_year': month_year,
        'month_number': month_number,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_raw': start_date_raw,
        'end_date_raw': end_date_raw,
        'machine_name': machine_name,
        'machine_operator': machine_operator,
        'item_id': item_id,
        'q': q,
    }


def _apply_daily_machine_filters(queryset, filters):
    if filters['month_year'] and filters['month_number']:
        queryset = queryset.filter(
            production_date__year=filters['month_year'],
            production_date__month=filters['month_number'],
        )

    if filters['start_date']:
        queryset = queryset.filter(production_date__gte=filters['start_date'])

    if filters['end_date']:
        queryset = queryset.filter(production_date__lte=filters['end_date'])

    if filters['machine_name']:
        queryset = queryset.filter(machine_name__icontains=filters['machine_name'])

    if filters['machine_operator']:
        queryset = queryset.filter(machine_operator__icontains=filters['machine_operator'])

    if filters['item_id']:
        queryset = queryset.filter(item_id=filters['item_id'])

    if filters.get('q'):
        queryset = queryset.filter(
            Q(machine_name__icontains=filters['q']) |
            Q(item__name__icontains=filters['q']) |
            Q(machine_operator__icontains=filters['q'])
        )

    return queryset


def _filtered_daily_machine_queryset(request):
    filters = _daily_machine_filters_from_request(request)
    queryset = DailyRunningMachine.objects.select_related('item').all()
    queryset = _apply_daily_machine_filters(queryset, filters)
    return queryset.order_by('-production_date', '-id'), filters


def _daily_machine_export_context(queryset, filters):
    selected_item = None
    if filters['item_id']:
        selected_item = Product.objects.filter(pk=filters['item_id']).first()

    if filters['month_year'] and filters['month_number']:
        month_label = datetime(filters['month_year'], filters['month_number'], 1).strftime('%B %Y')
    else:
        month_label = 'All Months'

    return {
        'machine_runs': queryset,
        'generated_at': timezone.localtime(),
        'month_label': month_label,
        'start_date': filters['start_date'],
        'end_date': filters['end_date'],
        'machine_name': filters['machine_name'],
        'machine_operator': filters['machine_operator'],
        'selected_item': selected_item,
        'total_rows': queryset.count(),
    }

def daily_machine_list(request):
    machine_runs_qs, filters = _filtered_daily_machine_queryset(request)
    paginator = Paginator(machine_runs_qs, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    products = Product.objects.select_related('category').all().order_by('name')
    machines = Machine.objects.all()
    work_assignments = DailyWorkAssignment.objects.order_by('-production_date')[:10]

    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'production/daily_running_machines/daily_machine_list.html', {
        'machine_runs': page_obj,
        'page_obj': page_obj,
        'products': products,
        'machines': machines,
        'work_assignments': work_assignments,
        'total_machine_runs': DailyRunningMachine.objects.count(),
        'filtered_count': machine_runs_qs.count(),
        'filters': filters,
        'query_string': query_params.urlencode(),
    })


def daily_machine_export_pdf(request):
    machine_runs_qs, filters = _filtered_daily_machine_queryset(request)
    context = _daily_machine_export_context(machine_runs_qs, filters)

    template = get_template('production/daily_running_machines/daily_machine_report_pdf.html')
    html = template.render(context)

    try:
        from xhtml2pdf import pisa

        response = HttpResponse(content_type='application/pdf')
        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="daily_machine_report_{timestamp}.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('PDF generation error', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/daily_running_machines/daily_machine_report_pdf.html', context)


def daily_machine_export_excel(request):
    machine_runs_qs, filters = _filtered_daily_machine_queryset(request)

    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse('Excel export dependency missing: openpyxl', status=500)

    context = _daily_machine_export_context(machine_runs_qs, filters)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Daily Machine Report'

    sheet.append(['Daily Machine Running Report'])
    sheet.append(['Generated At', context['generated_at'].strftime('%d %b %Y %I:%M %p')])
    sheet.append(['Month', context['month_label']])
    sheet.append(['Date From', context['start_date'].strftime('%d %b %Y') if context['start_date'] else 'All'])
    sheet.append(['Date To', context['end_date'].strftime('%d %b %Y') if context['end_date'] else 'All'])
    sheet.append(['Machine Name', context['machine_name'] or 'All'])
    sheet.append(['Operator', context['machine_operator'] or 'All'])
    sheet.append(['Item', context['selected_item'].name if context['selected_item'] else 'All'])
    sheet.append(['Total Rows', context['total_rows']])
    sheet.append([])
    sheet.append(['Production Date', 'Machine Name', 'Status', 'Item', 'Operator', 'Crusher Operator', 'Material Mixer', 'Extra Work Employee', 'Notes'])

    for run in machine_runs_qs:
        sheet.append([
            run.production_date.strftime('%d %b %Y'),
            run.machine_name,
            'Not Working' if run.machine_not_working else 'Working',
            run.item.name if run.item else '—',
            run.machine_operator or '—',
            run.crusher_operator or '—',
            run.material_mixer or '—',
            run.extra_work_employee or '—',
            run.notes or '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="daily_machine_report_{timestamp}.xlsx"'
    workbook.save(response)
    return response


def daily_machine_create(request):
    if request.method == 'POST':
        form = DailyRunningMachineForm(request.POST)
        if form.is_valid():
            run = form.save()
            log_activity(request, 'production', 'machine_run_created',
                         f"Machine run logged: {run.machine_name} | {run.production_date} | {run.item}",
                         reverse('daily_machine_list'), related_id=run.pk)
            messages.success(request, 'Daily machine production record created successfully.')
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, f'{error}')
    return redirect('daily_machine_list')


def daily_machine_update(request, pk):
    machine_run = get_object_or_404(DailyRunningMachine, pk=pk)
    if request.method == 'POST':
        form = DailyRunningMachineForm(request.POST, instance=machine_run)
        if form.is_valid():
            form.save()
            log_activity(request, 'production', 'machine_run_updated',
                         f"Machine run updated: {machine_run.machine_name} | {machine_run.production_date}",
                         reverse('daily_machine_list'))
            messages.success(request, 'Daily machine production record updated successfully.')
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, f'{error}')
    return redirect('daily_machine_list')


def daily_machine_delete(request, pk):
    machine_run = get_object_or_404(DailyRunningMachine, pk=pk)
    if request.method == 'POST':
        desc = f"Machine run deleted: {machine_run.machine_name} | {machine_run.production_date}"
        machine_run.delete()
        log_activity(request, 'production', 'machine_run_deleted', desc, reverse('daily_machine_list'))
        messages.success(request, 'Daily machine production record deleted successfully.')
    return redirect('daily_machine_list')


# ── Machine CRUD ─────────────────────────────────────────────

def machine_create(request):
    if request.method == 'POST':
        form = MachineForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Machine added successfully.')
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
    return redirect('daily_machine_list')


def machine_update(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        form = MachineForm(request.POST, instance=machine)
        if form.is_valid():
            form.save()
            messages.success(request, 'Machine updated successfully.')
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
    return redirect('daily_machine_list')


def machine_delete(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        machine.delete()
        messages.success(request, 'Machine deleted successfully.')
    return redirect('daily_machine_list')


# ── Daily Work Assignments ─────────────────────────────────────

def work_assignment_save(request):
    if request.method == 'POST':
        date_str = (request.POST.get('production_date') or '').strip()
        prod_date = parse_date(date_str) if date_str else None
        if prod_date:
            assignment, _ = DailyWorkAssignment.objects.get_or_create(production_date=prod_date)
            form = DailyWorkAssignmentForm(request.POST, instance=assignment)
            if form.is_valid():
                form.save()
                messages.success(request, 'Work assignment saved.')
            else:
                for errors in form.errors.values():
                    for error in errors:
                        messages.error(request, error)
        else:
            messages.error(request, 'A valid date is required for work assignment.')
    return redirect('daily_machine_list')


def work_assignment_delete(request, pk):
    assignment = get_object_or_404(DailyWorkAssignment, pk=pk)
    if request.method == 'POST':
        assignment.delete()
        messages.success(request, 'Work assignment deleted.')
    return redirect('daily_machine_list')


# ── Production Entries ─────────────────────────────────────────

def _production_entry_qs(request):
    qs = (
        ProductionEntry.objects
        .select_related('product', 'product__category', 'created_by')
        .order_by('-date', '-created_at')
    )
    date_from, date_to = resolve_date_range(request)
    product_id = request.GET.get('product', '').strip()
    category_id = request.GET.get('category', '').strip()

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if product_id:
        qs = qs.filter(product_id=product_id)
    if category_id:
        qs = qs.filter(product__category_id=category_id)

    return qs, date_from, date_to, product_id, category_id


def production_entry_list(request):
    qs, date_from, date_to, product_id, category_id = _production_entry_qs(request)

    totals = qs.aggregate(total_added=Sum('qty_added'))
    total_added = totals['total_added'] or Decimal('0')

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'production/production_entries/list.html', {
        'entries':      page_obj,
        'page_obj':     page_obj,
        'total_added':  total_added,
        'date_from':    date_from,
        'date_to':      date_to,
        'product_id':   product_id,
        'category_id':  category_id,
        'products':     Product.objects.select_related('category').order_by('name'),
        'categories':   ProductCategory.objects.order_by('name'),
        'today':        timezone.localdate(),
        'query_string': query_params.urlencode(),
    })


@transaction.atomic
def production_entry_create(request):
    if request.method != 'POST':
        return redirect('production_entry_list')

    try:
        product_id = request.POST.get('product')
        qty_raw    = request.POST.get('qty_added', '').strip()
        date_raw   = request.POST.get('date', '').strip()
        notes      = request.POST.get('notes', '').strip()

        product = Product.objects.select_for_update().get(pk=product_id)
        qty_added = Decimal(qty_raw)
        if qty_added <= 0:
            raise ValueError('Quantity must be greater than zero.')

        entry_date = parse_date(date_raw) if date_raw else timezone.localdate()

        qty_before = product.qty
        qty_after  = qty_before + qty_added

        ProductionEntry.objects.create(
            date=entry_date,
            product=product,
            qty_added=qty_added,
            qty_before=qty_before,
            qty_after=qty_after,
            notes=notes,
            created_by=request.user if request.user.is_authenticated else None,
        )

        product.qty = qty_after
        product.save(update_fields=['qty', 'updated_at'])

        log_activity(
            request, 'production', 'production_entry_added',
            f"Production: {product.name} +{qty_added} (total {qty_after})",
            reverse('production_entry_list'),
        )
        messages.success(request, f'Added {_fmt_qty(qty_added)} units to {product.name}. New total: {_fmt_qty(qty_after)}.')
    except Product.DoesNotExist:
        messages.error(request, 'Product not found.')
    except (InvalidOperation, ValueError) as exc:
        messages.error(request, str(exc) or 'Invalid quantity.')

    return redirect('production_entry_list')


@transaction.atomic
def production_entry_edit(request, pk):
    entry = get_object_or_404(ProductionEntry.objects.select_related('product'), pk=pk)
    if request.method != 'POST':
        return redirect('production_entry_list')

    try:
        qty_raw  = request.POST.get('qty_added', '').strip()
        date_raw = request.POST.get('date', '').strip()
        notes    = request.POST.get('notes', '').strip()

        new_qty_added = Decimal(qty_raw)
        if new_qty_added <= 0:
            raise ValueError('Quantity must be greater than zero.')

        entry_date = parse_date(date_raw) if date_raw else entry.date

        product = Product.objects.select_for_update().get(pk=entry.product_id)
        old_qty_added = entry.qty_added

        # Adjust live stock by exactly the difference; keep this entry's own
        # qty_before as the historical fact it was, and recompute qty_after
        # to stay internally consistent with the new qty_added.
        product.qty = product.qty - old_qty_added + new_qty_added
        product.save(update_fields=['qty', 'updated_at'])

        entry.qty_added = new_qty_added
        entry.qty_after = entry.qty_before + new_qty_added
        entry.date      = entry_date
        entry.notes     = notes
        entry.save(update_fields=['qty_added', 'qty_after', 'date', 'notes'])

        log_activity(
            request, 'production', 'production_entry_updated',
            f"Production entry updated: {product.name} {old_qty_added} → {new_qty_added} (stock now {_fmt_qty(product.qty)})",
            reverse('production_entry_list'),
        )
        messages.success(request, f'Production entry updated. {product.name} is now at {_fmt_qty(product.qty)}.')
    except ProductionEntry.DoesNotExist:
        messages.error(request, 'Production entry not found.')
    except (InvalidOperation, ValueError) as exc:
        messages.error(request, str(exc) or 'Invalid quantity.')

    return redirect('production_entry_list')


def production_entry_delete(request, pk):
    entry = get_object_or_404(ProductionEntry, pk=pk)
    if request.method != 'POST':
        return redirect('production_entry_list')

    with transaction.atomic():
        product = Product.objects.select_for_update().get(pk=entry.product_id)
        product.qty = product.qty - entry.qty_added
        product.save(update_fields=['qty', 'updated_at'])
        entry.delete()
        messages.success(request, f'Production entry deleted. Qty reversed.')

    return redirect('production_entry_list')


def production_entry_export_excel(request):
    qs, date_from, date_to, product_id, category_id = _production_entry_qs(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    totals = qs.aggregate(total_added=Sum('qty_added'))
    total_added = totals['total_added'] or Decimal('0')

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Production Entries'
    gen = timezone.localtime()

    ws.append(['Production Entry Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    if date_from or date_to:
        ws.append(['Period', f'{date_from or "—"} to {date_to or "—"}'])
    ws.append(['Total Units Added', float(total_added)])
    ws.append([])

    headers = ['#', 'Date', 'Category', 'Product', 'Qty Added', 'Qty Before', 'Qty After', 'Notes', 'Added By']
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, e in enumerate(qs, 1):
        ws.append([
            i,
            e.date.strftime('%d-%m-%Y'),
            e.product.category.name,
            e.product.name,
            float(e.qty_added),
            float(e.qty_before),
            float(e.qty_after),
            e.notes or '',
            e.created_by.get_full_name() or e.created_by.username if e.created_by else '',
        ])

    # totals row
    ws.append([])
    ws.append(['', '', '', 'TOTAL', float(total_added), '', '', '', ''])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)

    for col, w in zip('ABCDEFGHI', [4, 13, 18, 25, 12, 12, 12, 30, 16]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="production_entries_{ts}.xlsx"'
    wb.save(response)
    return response


def production_entry_export_pdf(request):
    qs, date_from, date_to, product_id, category_id = _production_entry_qs(request)

    totals = qs.aggregate(total_added=Sum('qty_added'))
    total_added = totals['total_added'] or Decimal('0')

    context = {
        'entries':     list(qs),
        'total_added': total_added,
        'date_from':   date_from,
        'date_to':     date_to,
        'generated_at': timezone.localtime(),
    }
    template = get_template('production/production_entries/pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="production_entries_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/production_entries/pdf.html', context)


# ═══════════════════════════════════════════════════════════════
# STOCK LEDGER — production additions + bill deductions, running balance
# ═══════════════════════════════════════════════════════════════

def _stock_ledger_products_qs(request):
    date_from, date_to = resolve_date_range(request)
    product_id  = request.GET.get('product', '').strip()
    category_id = request.GET.get('category', '').strip()

    products = Product.objects.select_related('category').order_by('name')
    if product_id:
        products = products.filter(pk=product_id)
    if category_id:
        products = products.filter(category_id=category_id)

    return products, date_from, date_to, product_id, category_id


def _product_ledger(product, date_from, date_to):
    """Chronological (production + sale) events for one product with running
    production-total and running-balance columns. Cancelled bills are excluded
    since their stock was already restored by bill_cancel."""
    entries_qs = ProductionEntry.objects.filter(product=product).order_by('date', 'id')
    items_qs = (
        BillItem.objects.filter(product=product)
        .exclude(bill__status=Bill.CANCELLED)
        .select_related('bill', 'bill__customer')
        .order_by('bill__bill_date', 'id')
    )

    # Opening balance = net of everything strictly before date_from, so a
    # filtered window still shows a true running balance, not one that
    # restarts at zero.
    opening_total   = Decimal('0')
    opening_balance = Decimal('0')
    if date_from:
        opening_total = entries_qs.filter(date__lt=date_from).aggregate(s=Sum('qty_added'))['s'] or Decimal('0')
        opening_sold  = items_qs.filter(bill__bill_date__lt=date_from).aggregate(s=Sum('quantity'))['s'] or Decimal('0')
        opening_balance = opening_total - opening_sold

    if date_from:
        entries_qs = entries_qs.filter(date__gte=date_from)
        items_qs   = items_qs.filter(bill__bill_date__gte=date_from)
    if date_to:
        entries_qs = entries_qs.filter(date__lte=date_to)
        items_qs   = items_qs.filter(bill__bill_date__lte=date_to)

    events = [(e.date, 0, e.id, 'production', e) for e in entries_qs]
    events += [(it.bill.bill_date, 1, it.id, 'sale', it) for it in items_qs]
    events.sort(key=lambda t: (t[0], t[1], t[2]))

    running_total   = opening_total
    running_balance = opening_balance
    rows = []
    for date_, _, _, kind, obj in events:
        if kind == 'production':
            running_total   += obj.qty_added
            running_balance += obj.qty_added
            rows.append({
                'date': date_, 'production': obj.qty_added, 'total': running_total,
                'sale': None, 'balance': running_balance, 'customer': '', 'bill_no': '', 'bill_id': None,
            })
        else:
            running_balance -= obj.quantity
            rows.append({
                'date': date_, 'production': None, 'total': None,
                'sale': obj.quantity, 'balance': running_balance,
                'customer': obj.bill.customer.name, 'bill_no': obj.bill.bill_number, 'bill_id': obj.bill_id,
            })
    return rows, opening_balance


def _stock_ledger_blocks(request):
    products, date_from, date_to, product_id, category_id = _stock_ledger_products_qs(request)
    blocks = []
    for product in products:
        rows, opening_balance = _product_ledger(product, date_from, date_to)
        if not rows:
            continue
        blocks.append({
            'product':          product,
            'rows':             rows,
            'opening_balance':  opening_balance,
            'closing_balance':  rows[-1]['balance'],
            'total_production': sum((r['production'] for r in rows if r['production'] is not None), Decimal('0')),
            'total_sale':       sum((r['sale'] for r in rows if r['sale'] is not None), Decimal('0')),
        })
    return blocks, date_from, date_to, product_id, category_id


def stock_ledger(request):
    blocks, date_from, date_to, product_id, category_id = _stock_ledger_blocks(request)

    # Paginate by product block (not by row) — each card keeps its own
    # complete, uninterrupted running balance. Exports always use the
    # full, unpaginated `blocks` list.
    paginator = Paginator(blocks, 10)
    page_obj  = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'production/stock_ledger.html', {
        'blocks':       page_obj,
        'page_obj':     page_obj,
        'date_from':    date_from,
        'date_to':      date_to,
        'product_id':   product_id,
        'category_id':  category_id,
        'products':     Product.objects.select_related('category').order_by('name'),
        'categories':   ProductCategory.objects.order_by('name'),
        'query_string': query_params.urlencode(),
    })


def stock_ledger_export_excel(request):
    blocks, date_from, date_to, product_id, category_id = _stock_ledger_blocks(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Stock Ledger'
    gen = timezone.localtime()

    ws.append(['Stock Ledger Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    if date_from or date_to:
        ws.append(['Period', f'{date_from or "—"} to {date_to or "—"}'])
    ws.append([])

    headers = ['Date', 'Production', 'Total', 'Sale', 'Balance', 'Customer', 'Bill No.']
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    pfill = PatternFill('solid', fgColor='F3F4F6')

    for block in blocks:
        product = block['product']
        label   = product.name + (f' ({product.size})' if product.size else '')
        ws.append([f'{label} — {product.category.name}'])
        prow = ws.max_row
        ws.cell(row=prow, column=1).font = Font(bold=True, size=11)
        ws.cell(row=prow, column=1).fill = pfill
        ws.merge_cells(start_row=prow, start_column=1, end_row=prow, end_column=len(headers))

        ws.append(headers)
        hrow = ws.max_row
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=hrow, column=ci)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center')

        for r in block['rows']:
            ws.append([
                r['date'].strftime('%d-%m-%Y'),
                float(r['production']) if r['production'] is not None else None,
                float(r['total'])      if r['total']      is not None else None,
                float(r['sale'])       if r['sale']       is not None else None,
                float(r['balance']),
                r['customer'],
                r['bill_no'],
            ])
        ws.append([])

    if not blocks:
        ws.append(['No stock movement found for the selected filters.'])

    for col, w in zip('ABCDEFG', [13, 12, 12, 12, 12, 20, 16]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="stock_ledger_{ts}.xlsx"'
    wb.save(response)
    return response


def stock_ledger_export_pdf(request):
    blocks, date_from, date_to, product_id, category_id = _stock_ledger_blocks(request)
    context = {
        'blocks':       blocks,
        'date_from':    date_from,
        'date_to':      date_to,
        'generated_at': timezone.localtime(),
    }
    template = get_template('production/stock_ledger_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="stock_ledger_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/stock_ledger_pdf.html', context)


# ═══════════════════════════════════════════════════════════════
# ALL PRODUCTION REPORT — live snapshot of every product's available qty
# ═══════════════════════════════════════════════════════════════

def _all_production_report_qs(request):
    q = (request.GET.get('q') or '').strip()
    category_id = (request.GET.get('category') or '').strip()

    qs = Product.objects.select_related('category').order_by('category__name', 'name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(category__name__icontains=q) | Q(size__icontains=q))
    if category_id:
        qs = qs.filter(category_id=category_id)

    return qs, q, category_id


def all_production_report(request):
    products_qs, q, category_id = _all_production_report_qs(request)
    total_qty = products_qs.aggregate(s=Sum('qty'))['s'] or Decimal('0')

    paginator = Paginator(products_qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'production/all_production_report.html', {
        'products':       page_obj,
        'page_obj':       page_obj,
        'total_products': products_qs.count(),
        'total_qty':      total_qty,
        'categories':     ProductCategory.objects.order_by('name'),
        'q':              q,
        'category_id':    category_id,
        'today':          timezone.localdate(),
        'query_string':   query_params.urlencode(),
    })


def all_production_report_export_excel(request):
    products_qs, q, category_id = _all_production_report_qs(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Production Report'
    gen   = timezone.localtime()
    today = timezone.localdate()

    ws.append(['All Production Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['As Of', today.strftime('%d %b %Y')])
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Products', products_qs.count()])
    ws.append([])

    headers = ['#', 'Category', 'Product', 'Size', 'Available Qty']
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(products_qs, 1):
        ws.append([i, p.category.name, p.name, p.size or '', float(p.qty)])

    for col, w in zip('ABCDE', [4, 20, 28, 14, 14]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="all_production_report_{ts}.xlsx"'
    wb.save(response)
    return response


def all_production_report_export_pdf(request):
    products_qs, q, category_id = _all_production_report_qs(request)

    context = {
        'products':       products_qs,
        'total_products': products_qs.count(),
        'total_qty':      products_qs.aggregate(s=Sum('qty'))['s'] or Decimal('0'),
        'today':          timezone.localdate(),
        'generated_at':   timezone.localtime(),
    }
    template = get_template('production/all_production_report_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="all_production_report_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/all_production_report_pdf.html', context)


# ═══════════════════════════════════════════════════════════════
# TARGET LOG SYSTEM
# ═══════════════════════════════════════════════════════════════

def production_item_create(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = ProductionItemForm(request.POST)
    if form.is_valid():
        item = form.save()
        return JsonResponse({'ok': True, 'id': item.pk, 'name': item.name, 'hourly_qty': str(item.hourly_qty)})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def production_item_update(request, pk):
    item = get_object_or_404(ProductionItem, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = ProductionItemForm(request.POST, instance=item)
    if form.is_valid():
        item = form.save()
        return JsonResponse({'ok': True, 'name': item.name, 'hourly_qty': str(item.hourly_qty)})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def production_item_delete(request, pk):
    item = get_object_or_404(ProductionItem, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        item.delete()
        return JsonResponse({'ok': True})
    except ProtectedError:
        return JsonResponse({'ok': False, 'error': 'Item has existing target log entries and cannot be deleted.'}, status=400)


def shift_template_create(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = ShiftTemplateForm(request.POST)
    if form.is_valid():
        tpl = form.save()
        return JsonResponse({'ok': True, 'id': tpl.pk, 'name': tpl.name, 'duration_display': tpl.duration_display, 'duration_hours': tpl.duration_hours})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def shift_template_update(request, pk):
    tpl = get_object_or_404(ShiftTemplate, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = ShiftTemplateForm(request.POST, instance=tpl)
    if form.is_valid():
        tpl = form.save()
        return JsonResponse({'ok': True, 'name': tpl.name, 'duration_display': tpl.duration_display, 'duration_hours': tpl.duration_hours})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def shift_template_delete(request, pk):
    tpl = get_object_or_404(ShiftTemplate, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        tpl.delete()
        return JsonResponse({'ok': True})
    except ProtectedError:
        return JsonResponse({'ok': False, 'error': 'Template has existing target log entries.'}, status=400)


def target_log_list(request):
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    employee_id = request.GET.get('employee', '')
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
        month = f'{year:04d}-{mon:02d}'
    qs = TargetLog.objects.select_related('employee', 'item', 'shift_template').filter(
        date__year=year, date__month=mon,
    )
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    qs = qs.order_by('-date', '-created_at')
    employees = Employee.objects.filter(active=True).order_by('name')
    items = ProductionItem.objects.filter(active=True).order_by('name')
    all_items = ProductionItem.objects.all().order_by('name')
    templates = ShiftTemplate.objects.all().order_by('name')
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'production/target_logs/list.html', {
        'page_obj': page_obj, 'month': month, 'employee_id': employee_id,
        'employees': employees, 'items': items, 'all_items': all_items,
        'templates': templates, 'today': today.isoformat(),
    })


def target_log_create(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = TargetLogForm(request.POST)
    if form.is_valid():
        log = form.save()
        return JsonResponse({
            'ok': True, 'id': log.pk,
            'date': str(log.date), 'employee': log.employee.name,
            'machine_name': log.machine_name, 'item': log.item.name,
            'shift': log.shift_template.name if log.shift_template else '-',
            'target_qty': log.target_qty, 'actual_qty': log.actual_qty,
            'point': log.point, 'downtime_minutes': str(log.downtime_minutes),
            'downtime_reason': log.downtime_reason, 'remarks': log.remarks,
        })
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def target_log_update_inline(request, pk):
    log = get_object_or_404(TargetLog, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    field = request.POST.get('field')
    value = request.POST.get('value', '').strip()
    allowed = {'date', 'machine_name', 'actual_qty', 'target_qty', 'downtime_minutes',
               'downtime_reason', 'point', 'remarks', 'employee', 'item', 'shift_template',
               'cavity', 'cycle_time_seconds'}
    if field not in allowed:
        return JsonResponse({'ok': False, 'error': 'Invalid field'}, status=400)
    try:
        if field == 'employee':
            log.employee = get_object_or_404(Employee, pk=int(value))
        elif field == 'item':
            log.item = get_object_or_404(ProductionItem, pk=int(value))
        elif field == 'shift_template':
            log.shift_template = get_object_or_404(ShiftTemplate, pk=int(value)) if value else None
        elif field in ('actual_qty', 'target_qty', 'cavity', 'point'):
            setattr(log, field, int(value))
        elif field == 'cycle_time_seconds':
            setattr(log, field, int(value) if value else None)
        elif field == 'downtime_minutes':
            setattr(log, field, Decimal(value))
        elif field == 'date':
            from django.utils.dateparse import parse_date as _pd
            log.date = _pd(value)
        else:
            setattr(log, field, value)
        log.save()
    except (ValueError, InvalidOperation) as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    disp_map = {'employee': log.employee.name, 'item': log.item.name,
                'shift_template': log.shift_template.name if log.shift_template else '-'}
    return JsonResponse({
        'ok': True,
        'display': disp_map.get(field, str(getattr(log, field))),
        'point': log.point, 'auto_point': log.auto_point,
    })


def target_log_delete(request, pk):
    log = get_object_or_404(TargetLog, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    log.delete()
    return JsonResponse({'ok': True})


def target_log_calc_api(request):
    item_id = request.GET.get('item')
    tpl_id = request.GET.get('shift_template')
    if not item_id or not tpl_id:
        return JsonResponse({'target_qty': 0})
    try:
        item = ProductionItem.objects.get(pk=item_id)
        tpl = ShiftTemplate.objects.get(pk=tpl_id)
        return JsonResponse({'target_qty': TargetLog.calc_target(item, tpl), 'duration_display': tpl.duration_display})
    except (ProductionItem.DoesNotExist, ShiftTemplate.DoesNotExist):
        return JsonResponse({'target_qty': 0})


def target_log_points_report(request):
    from django.db.models import Sum, Count
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
        month = f'{year:04d}-{mon:02d}'
    qs = TargetLog.objects.filter(date__year=year, date__month=mon)
    rows_raw = list(qs.values('employee__id', 'employee__name')
        .annotate(total_entries=Count('id'), total_points=Sum('point'))
        .order_by('employee__name'))
    rows = [{'employee_id': r['employee__id'], 'employee_name': r['employee__name'],
              'total_entries': r['total_entries'], 'total_points': r['total_points'] or 0,
              'not_achieved': r['total_entries'] - (r['total_points'] or 0)} for r in rows_raw]
    return render(request, 'production/target_logs/points_report.html', {
        'rows': rows, 'month': month, 'year': year, 'mon': mon,
    })


def target_log_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    employee_id = request.GET.get('employee', '')
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = TargetLog.objects.select_related('employee', 'item', 'shift_template').filter(
        date__year=year, date__month=mon)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    qs = qs.order_by('date', 'employee__name')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Target Logs'
    ws.append([f'Target Log Report — {month}'])
    ws.append([])
    headers = ['Date', 'Employee', 'Machine', 'Item', 'Shift', 'Target', 'Actual', 'Downtime (min)', 'Reason', 'Point', 'Remarks']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1D4ED8')
        cell.alignment = Alignment(horizontal='center')
    for log in qs:
        ws.append([str(log.date), log.employee.name, log.machine_name, log.item.name,
                   log.shift_template.name if log.shift_template else '',
                   log.target_qty, log.actual_qty, float(log.downtime_minutes),
                   log.downtime_reason, log.point, log.remarks])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="target_logs_{ts}.xlsx"'
    wb.save(response)
    return response


def target_log_export_pdf(request):
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    employee_id = request.GET.get('employee', '')
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = TargetLog.objects.select_related('employee', 'item', 'shift_template').filter(
        date__year=year, date__month=mon)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    qs = qs.order_by('date', 'employee__name')
    context = {'logs': list(qs), 'month': month, 'generated_at': timezone.localtime()}
    template = get_template('production/target_logs/pdf.html')
    html = template.render(context)
    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="target_logs_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/target_logs/pdf.html', context)


def target_log_points_pdf(request):
    from django.db.models import Sum, Count
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = TargetLog.objects.filter(date__year=year, date__month=mon)
    rows_raw = list(qs.values('employee__id', 'employee__name')
        .annotate(total_entries=Count('id'), total_points=Sum('point'))
        .order_by('employee__name'))
    rows = [{'employee_name': r['employee__name'], 'total_entries': r['total_entries'],
              'total_points': r['total_points'] or 0,
              'not_achieved': r['total_entries'] - (r['total_points'] or 0)} for r in rows_raw]
    context = {'rows': rows, 'month': month, 'generated_at': timezone.localtime()}
    template = get_template('production/target_logs/points_pdf.html')
    html = template.render(context)
    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="target_points_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/target_logs/points_pdf.html', context)


def target_log_points_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.db.models import Sum, Count
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = TargetLog.objects.filter(date__year=year, date__month=mon)
    rows_raw = list(qs.values('employee__name')
        .annotate(total_entries=Count('id'), total_points=Sum('point'))
        .order_by('employee__name'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Points Report'
    ws.append([f'Target Achievement Points — {month}'])
    ws.append([])
    headers = ['Employee', 'Total Entries', 'Points (Achieved)', 'Not Achieved', 'Achievement %']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1D4ED8')
        cell.alignment = Alignment(horizontal='center')
    for r in rows_raw:
        total = r['total_entries']
        pts = r['total_points'] or 0
        pct = round(pts / total * 100, 1) if total else 0
        ws.append([r['employee__name'], total, pts, total - pts, f'{pct}%'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="target_points_{ts}.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════
# SHIFT PRODUCTION LOG SYSTEM
# ═══════════════════════════════════════════════════════════════

def employee_create(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = EmployeeForm(request.POST)
    if form.is_valid():
        emp = form.save()
        return JsonResponse({'ok': True, 'id': emp.pk, 'name': emp.name})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def employee_update(request, pk):
    emp = get_object_or_404(Employee, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    form = EmployeeForm(request.POST, instance=emp)
    if form.is_valid():
        emp = form.save()
        return JsonResponse({'ok': True, 'name': emp.name})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


def employee_delete(request, pk):
    emp = get_object_or_404(Employee, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        emp.delete()
        return JsonResponse({'ok': True})
    except ProtectedError:
        return JsonResponse({'ok': False, 'error': 'Employee has existing log entries.'}, status=400)


def shift_log_list(request):
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    employee_id = request.GET.get('employee', '')
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
        month = f'{year:04d}-{mon:02d}'
    qs = ShiftProductionLog.objects.select_related('employee').filter(
        shift_start__year=year, shift_start__month=mon)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    qs = qs.order_by('-shift_start')
    employees = Employee.objects.filter(active=True).order_by('name')
    all_employees = Employee.objects.all().order_by('name')
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'production/shift_logs/list.html', {
        'page_obj': page_obj, 'month': month, 'employee_id': employee_id,
        'employees': employees, 'all_employees': all_employees,
    })


def shift_log_create(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    import json
    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST.dict()
    try:
        from datetime import datetime as dt
        shift_start = dt.fromisoformat(data['shift_start'])
        shift_end = dt.fromisoformat(data['shift_end'])
        employee = get_object_or_404(Employee, pk=int(data['employee']))
        cavity = int(data['cavity'])
        cycle_time = int(data['cycle_time_seconds'])
        actual_qty = int(data['actual_qty'])
        target_qty = ShiftProductionLog.calc_target(shift_start, shift_end, cycle_time, cavity)
        auto_pt = 1 if actual_qty < target_qty else 0
        log = ShiftProductionLog.objects.create(
            shift_start=shift_start, shift_end=shift_end,
            machine_no=data['machine_no'], item_name=data['item_name'],
            cavity=cavity, cycle_time_seconds=cycle_time,
            target_qty=target_qty, actual_qty=actual_qty,
            downtime_minutes=Decimal(str(data.get('downtime_minutes', '0') or '0')),
            downtime_reason=data.get('downtime_reason', ''),
            point=auto_pt,
            employee=employee, remarks=data.get('remarks', ''),
        )
        return JsonResponse({'ok': True, 'id': log.pk, 'target_qty': log.target_qty,
                             'point': log.point, 'penalty_points': log.penalty_points,
                             'employee': log.employee.name,
                             'shift_start': str(log.shift_start), 'shift_end': str(log.shift_end)})
    except (ValueError, InvalidOperation, KeyError) as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


def shift_log_update_inline(request, pk):
    log = get_object_or_404(ShiftProductionLog, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    field = request.POST.get('field')
    value = request.POST.get('value', '').strip()
    allowed = {'shift_start', 'shift_end', 'machine_no', 'item_name', 'cavity',
               'cycle_time_seconds', 'actual_qty', 'target_qty', 'downtime_minutes',
               'downtime_reason', 'point', 'employee', 'remarks'}
    if field not in allowed:
        return JsonResponse({'ok': False, 'error': 'Invalid field'}, status=400)
    try:
        if field == 'employee':
            log.employee = get_object_or_404(Employee, pk=int(value))
        elif field in ('cavity', 'cycle_time_seconds', 'actual_qty', 'target_qty', 'point'):
            setattr(log, field, int(value))
        elif field == 'downtime_minutes':
            setattr(log, field, Decimal(value))
        elif field in ('shift_start', 'shift_end'):
            from datetime import datetime as dt
            setattr(log, field, dt.fromisoformat(value))
        else:
            setattr(log, field, value)
        log.save()
    except (ValueError, InvalidOperation) as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': True, 'point': log.point, 'penalty_points': log.penalty_points,
                         'display': log.employee.name if field == 'employee' else value})


def shift_log_delete(request, pk):
    log = get_object_or_404(ShiftProductionLog, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    log.delete()
    return JsonResponse({'ok': True})


def shift_log_points_report(request):
    from django.db.models import Count
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
        month = f'{year:04d}-{mon:02d}'
    qs = ShiftProductionLog.objects.filter(shift_start__year=year, shift_start__month=mon)
    rows_raw = list(qs.values('employee__id', 'employee__name')
        .annotate(total_entries=Count('id')).order_by('employee__name'))
    result = []
    for r in rows_raw:
        emp_qs = qs.filter(employee_id=r['employee__id'])
        penalty = sum(1 for log in emp_qs if log.actual_qty < log.target_qty)
        result.append({'employee_id': r['employee__id'], 'employee_name': r['employee__name'],
                       'total_entries': r['total_entries'], 'penalty_points': penalty,
                       'on_target': r['total_entries'] - penalty})
    return render(request, 'production/shift_logs/points_report.html', {
        'rows': result, 'month': month, 'year': year, 'mon': mon,
    })


def shift_log_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    employee_id = request.GET.get('employee', '')
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = ShiftProductionLog.objects.select_related('employee').filter(
        shift_start__year=year, shift_start__month=mon)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    qs = qs.order_by('shift_start')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Shift Logs'
    ws.append([f'Shift Production Log — {month}'])
    ws.append([])
    headers = ['Shift Start', 'Shift End', 'Employee', 'Machine', 'Item', 'Cavity',
               'Cycle (s)', 'Target', 'Actual', 'Downtime (min)', 'Reason', 'Penalty Pts', 'Remarks']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1D4ED8')
        cell.alignment = Alignment(horizontal='center')
    for log in qs:
        ws.append([str(log.shift_start), str(log.shift_end), log.employee.name, log.machine_no,
                   log.item_name, log.cavity, log.cycle_time_seconds, log.target_qty, log.actual_qty,
                   float(log.downtime_minutes), log.downtime_reason, log.penalty_points, log.remarks])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="shift_logs_{ts}.xlsx"'
    wb.save(response)
    return response


def shift_log_export_pdf(request):
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    employee_id = request.GET.get('employee', '')
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = ShiftProductionLog.objects.select_related('employee').filter(
        shift_start__year=year, shift_start__month=mon)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    qs = qs.order_by('shift_start')
    context = {'logs': list(qs), 'month': month, 'generated_at': timezone.localtime()}
    template = get_template('production/shift_logs/pdf.html')
    html = template.render(context)
    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="shift_logs_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/shift_logs/pdf.html', context)


def shift_log_points_pdf(request):
    from django.db.models import Count
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = ShiftProductionLog.objects.filter(shift_start__year=year, shift_start__month=mon)
    rows_raw = list(qs.values('employee__id', 'employee__name')
        .annotate(total_entries=Count('id')).order_by('employee__name'))
    result = [{'employee_name': r['employee__name'], 'total_entries': r['total_entries'],
               'penalty_points': sum(1 for l in qs.filter(employee_id=r['employee__id']) if l.actual_qty < l.target_qty),
               'on_target': r['total_entries'] - sum(1 for l in qs.filter(employee_id=r['employee__id']) if l.actual_qty < l.target_qty)}
              for r in rows_raw]
    context = {'rows': result, 'month': month, 'generated_at': timezone.localtime()}
    template = get_template('production/shift_logs/points_pdf.html')
    html = template.render(context)
    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="shift_points_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'production/shift_logs/points_pdf.html', context)


def shift_log_points_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.db.models import Count
    today = timezone.localdate()
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year, mon = today.year, today.month
    qs = ShiftProductionLog.objects.filter(shift_start__year=year, shift_start__month=mon)
    rows_raw = list(qs.values('employee__id', 'employee__name')
        .annotate(total_entries=Count('id')).order_by('employee__name'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Points Report'
    ws.append([f'Shift Log Penalty Points — {month}'])
    ws.append([])
    headers = ['Employee', 'Total Entries', 'On Target', 'Penalty Points', 'Hit Rate %']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1D4ED8')
        cell.alignment = Alignment(horizontal='center')
    for r in rows_raw:
        emp_qs = qs.filter(employee_id=r['employee__id'])
        penalty = sum(1 for l in emp_qs if l.actual_qty < l.target_qty)
        on_target = r['total_entries'] - penalty
        pct = round(on_target / r['total_entries'] * 100, 1) if r['total_entries'] else 0
        ws.append([r['employee__name'], r['total_entries'], on_target, penalty, f'{pct}%'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="shift_points_{ts}.xlsx"'
    wb.save(response)
    return response
