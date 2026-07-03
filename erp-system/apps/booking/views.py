import io
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone

from apps.billing.models import Bill, BillItem
from apps.customers.models import Customer, CustomerProductPrice
from apps.production.models import Product
from .models import BookingOrder, BookingItem
from erp.utils import log_activity

# ─────────────────────────── helpers ──────────────────────────────────────────

_TWO = Decimal('0.01')


def _fmt_qty(q):
    return f'{q:.2f}'.rstrip('0').rstrip('.')


def _d(value, default='0'):
    try:
        return Decimal(str(value or default)).quantize(_TWO, rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError):
        return Decimal(default)


def _filter_qs_str(request):
    """Return url-encoded query string without the 'page' param."""
    params = request.GET.copy()
    params.pop('page', None)
    return params.urlencode()


def _apply_booking_filters(qs, q, customer, date_from, date_to, status):
    if q:
        qs = qs.filter(
            Q(booking_number__icontains=q) | Q(customer__name__icontains=q)
        )
    if customer:
        qs = qs.filter(customer__name__icontains=customer)
    if date_from:
        qs = qs.filter(booking_date__gte=date_from)
    if date_to:
        qs = qs.filter(booking_date__lte=date_to)
    if status:
        qs = qs.filter(status=status)
    return qs


# ─────────────────────────── booking list ─────────────────────────────────────

def booking_list(request):
    q         = request.GET.get('q', '').strip()
    customer  = request.GET.get('customer', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    status    = request.GET.get('status', '').strip()

    qs = BookingOrder.objects.select_related('customer', 'bill').order_by('-booking_date', '-id')
    qs = _apply_booking_filters(qs, q, customer, date_from, date_to, status)

    agg           = qs.aggregate(total=Sum('total_amount'))
    total_amount  = agg['total'] or Decimal('0')
    total_records = qs.count()
    pending_count   = qs.filter(status=BookingOrder.PENDING).count()
    confirmed_count = qs.filter(status=BookingOrder.CONFIRMED).count()

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    customers = Customer.objects.order_by('name')

    return render(request, 'booking/booking_list.html', {
        'page_obj':       page_obj,
        'bookings':       page_obj.object_list,
        'total_records':  total_records,
        'total_amount':   total_amount,
        'pending_count':  pending_count,
        'confirmed_count': confirmed_count,
        'customers':      customers,
        'filter_qs':      _filter_qs_str(request),
        # filter values for form repopulation
        'q':         q,
        'customer':  customer,
        'date_from': date_from,
        'date_to':   date_to,
        'status':    status,
        'status_choices': BookingOrder.STATUS_CHOICES,
    })


# ─────────────────────────── print selected bookings ─────────────────────────

def booking_print_selected(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
    else:
        ids = request.GET.get('ids', '').split(',')

    ids = [i for i in ids if i.strip().isdigit()]
    if not ids:
        messages.warning(request, 'No bookings selected.')
        return redirect('booking:booking_list')

    bookings = (
        BookingOrder.objects
        .filter(pk__in=ids)
        .select_related('customer')
        .prefetch_related('items__product')
        .order_by('booking_date', 'id')
    )
    return render(request, 'booking/print_selected.html', {
        'bookings':     bookings,
        'generated_at': timezone.now(),
    })


# ─────────────────────────── booking create ───────────────────────────────────

def booking_create(request):
    customers = Customer.objects.order_by('name')

    if request.method == 'POST':
        try:
            booking = _create_booking(request)
            from django.urls import reverse
            log_activity(
                request, 'booking', 'booking_created',
                f"Booking created: {booking.booking_number} | {booking.customer.name} | Rs. {booking.total_amount}",
                reverse('booking:booking_detail', kwargs={'pk': booking.pk}),
                related_id=booking.pk,
            )
            messages.success(request, f'Booking {booking.booking_number} created successfully.')
            return redirect('booking:booking_detail', pk=booking.pk)
        except (ValueError, Product.DoesNotExist, Customer.DoesNotExist) as exc:
            messages.error(request, str(exc))

    return render(request, 'booking/booking_create.html', {
        'customers': customers,
        'today':     timezone.localdate().isoformat(),
    })


@transaction.atomic
def _create_booking(request):
    POST = request.POST

    # customer
    customer_id = POST.get('customer_id', '').strip()
    if not customer_id:
        raise ValueError('Please select a customer.')
    customer = Customer.objects.get(pk=customer_id)

    # booking date
    date_str = POST.get('booking_date', '') or timezone.localdate().isoformat()
    try:
        booking_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        booking_date = timezone.localdate()

    # order sending date (optional)
    osd_str = POST.get('order_sending_date', '').strip()
    order_sending_date = None
    if osd_str:
        try:
            order_sending_date = datetime.strptime(osd_str, '%Y-%m-%d').date()
        except ValueError:
            order_sending_date = None

    notes = POST.get('notes', '').strip()

    # items
    try:
        items_raw = json.loads(POST.get('items_json', '[]') or '[]')
    except json.JSONDecodeError:
        raise ValueError('Invalid items data — please refresh and try again.')
    if not items_raw:
        raise ValueError('Please add at least one product to the booking.')

    subtotal       = Decimal('0')
    total_discount = Decimal('0')
    validated      = []

    for raw in items_raw:
        pid = raw.get('product_id')
        product = Product.objects.get(pk=pid)

        qty        = _d(raw.get('quantity'), '0')
        unit_price = _d(raw.get('unit_price'), '0')
        disc_pct   = _d(raw.get('discount_percent'), '0')

        if qty <= 0:
            raise ValueError(f'Quantity for "{product.name}" must be > 0.')
        if unit_price < 0:
            raise ValueError(f'Unit price for "{product.name}" cannot be negative.')
        if not (0 <= disc_pct <= 100):
            raise ValueError(f'Discount for "{product.name}" must be 0–100%.')

        disc_amt   = (qty * unit_price * disc_pct / 100).quantize(_TWO, rounding=ROUND_HALF_UP)
        line_total = (qty * unit_price - disc_amt).quantize(_TWO, rounding=ROUND_HALF_UP)

        subtotal       += (qty * unit_price).quantize(_TWO, rounding=ROUND_HALF_UP)
        total_discount += disc_amt
        validated.append({
            'product':          product,
            'quantity':         qty,
            'unit_price':       unit_price,
            'discount_percent': disc_pct,
            'discount_amount':  disc_amt,
            'line_total':       line_total,
        })

    total_amount = (subtotal - total_discount).quantize(_TWO, rounding=ROUND_HALF_UP)

    booking = BookingOrder.objects.create(
        customer           = customer,
        booking_date       = booking_date,
        order_sending_date = order_sending_date,
        status             = BookingOrder.PENDING,
        subtotal           = subtotal,
        discount_amount    = total_discount,
        total_amount       = total_amount,
        notes              = notes,
        created_by         = request.user if request.user.is_authenticated else None,
    )

    for v in validated:
        BookingItem.objects.create(
            booking          = booking,
            product          = v['product'],
            quantity         = v['quantity'],
            unit_price       = v['unit_price'],
            discount_percent = v['discount_percent'],
            discount_amount  = v['discount_amount'],
            line_total       = v['line_total'],
        )

    return booking


# ─────────────────────────── booking detail ───────────────────────────────────

def booking_detail(request, pk):
    booking = get_object_or_404(
        BookingOrder.objects.select_related('customer', 'created_by', 'bill')
                            .prefetch_related('items__product__category'),
        pk=pk,
    )
    return render(request, 'booking/booking_detail.html', {'booking': booking})


# ─────────────────────────── booking confirm ──────────────────────────────────

def booking_confirm(request, pk):
    booking = get_object_or_404(
        BookingOrder.objects.select_related('customer').prefetch_related('items__product'),
        pk=pk,
    )

    if booking.status != BookingOrder.PENDING:
        messages.error(request, 'Only pending bookings can be confirmed.')
        return redirect('booking:booking_detail', pk=pk)

    if request.method == 'POST':
        try:
            bill = _confirm_booking(request, booking)
            from django.urls import reverse
            log_activity(
                request, 'booking', 'booking_confirmed',
                f"Booking confirmed: {booking.booking_number} | {booking.customer.name} | Bill {bill.bill_number}",
                reverse('booking:booking_detail', kwargs={'pk': booking.pk}),
            )
            messages.success(
                request,
                f'Booking {booking.booking_number} confirmed. Bill {bill.bill_number} created and is pending payment.'
            )
            return redirect('billing:pending_bills')
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect('booking:booking_detail', pk=pk)

    return render(request, 'booking/booking_confirm.html', {'booking': booking})


@transaction.atomic
def _confirm_booking(request, booking):
    # lock the booking row
    booking = BookingOrder.objects.select_for_update().get(pk=booking.pk)
    if booking.status != BookingOrder.PENDING:
        raise ValueError('Booking is no longer pending.')

    # lock customer
    customer = Customer.objects.select_for_update().get(pk=booking.customer_id)

    items = booking.items.select_related('product').all()

    # stock check
    for item in items:
        product = Product.objects.select_for_update().get(pk=item.product_id)
        if product.qty < item.quantity:
            raise ValueError(
                f'Insufficient stock for "{product.name}". '
                f'Available: {_fmt_qty(product.qty)}, Requested: {_fmt_qty(item.quantity)}.'
            )

    # create bill
    bill = Bill.objects.create(
        customer        = customer,
        bill_date       = timezone.localdate(),
        status          = Bill.PENDING,
        payment_method  = Bill.PAY_LATER,
        subtotal        = booking.subtotal,
        discount_amount = booking.discount_amount,
        total_amount    = booking.total_amount,
        balance_used    = Decimal('0'),
        amount_paid     = Decimal('0'),
        amount_due      = booking.total_amount,
        notes           = booking.notes,
        created_by      = request.user if request.user.is_authenticated else None,
    )

    # create bill items and deduct inventory
    for item in items:
        BillItem.objects.create(
            bill             = bill,
            product          = item.product,
            quantity         = item.quantity,
            unit_price       = item.unit_price,
            discount_percent = item.discount_percent,
            discount_amount  = item.discount_amount,
            line_total       = item.line_total,
        )
        Product.objects.filter(pk=item.product_id).update(qty=F('qty') - item.quantity)

    # update booking
    booking.status = BookingOrder.CONFIRMED
    booking.bill   = bill
    booking.save(update_fields=['status', 'bill', 'updated_at'])

    return bill


# ─────────────────────────── booking cancel ───────────────────────────────────

def booking_cancel(request, pk):
    booking = get_object_or_404(BookingOrder, pk=pk)

    if booking.status != BookingOrder.PENDING:
        messages.error(request, 'Only pending bookings can be cancelled.')
        return redirect('booking:booking_detail', pk=pk)

    if request.method == 'POST':
        booking.status = BookingOrder.CANCELLED
        booking.save(update_fields=['status', 'updated_at'])
        from django.urls import reverse
        log_activity(
            request, 'booking', 'booking_cancelled',
            f"Booking cancelled: {booking.booking_number} | {booking.customer.name}",
            reverse('booking:booking_detail', kwargs={'pk': booking.pk}),
        )
        messages.success(request, f'Booking {booking.booking_number} has been cancelled.')
        return redirect('booking:booking_list')

    return render(request, 'booking/booking_cancel.html', {'booking': booking})


# ─────────────────────────── quotation PDF ────────────────────────────────────

def booking_quotation_pdf(request, pk):
    booking = get_object_or_404(
        BookingOrder.objects.select_related('customer').prefetch_related('items__product__category'),
        pk=pk,
    )
    from xhtml2pdf import pisa

    template   = get_template('booking/quotation_pdf.html')
    html       = template.render({
        'booking':       booking,
        'generated_at':  timezone.now(),
    })
    response = HttpResponse(content_type='application/pdf')
    filename = f"quotation_{booking.booking_number}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    pisa.CreatePDF(html, dest=response)
    return response


# ─────────────────────────── quotation Excel ──────────────────────────────────

def booking_quotation_excel(request, pk):
    booking = get_object_or_404(
        BookingOrder.objects.select_related('customer').prefetch_related('items__product__category'),
        pk=pk,
    )
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Quotation'

    blue_fill   = PatternFill('solid', fgColor='1D4ED8')
    white_bold  = Font(color='FFFFFF', bold=True, size=10)
    bold_font   = Font(bold=True, size=10)
    header_font = Font(bold=True, size=14)
    normal_font = Font(size=10)
    right_align = Alignment(horizontal='right')
    center_align= Alignment(horizontal='center')

    thin_side = Side(style='thin', color='E5E7EB')
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'QUOTATION — SENOVKA ERP'
    ws['A1'].font = Font(bold=True, size=16, color='1D4ED8')
    ws['A1'].alignment = center_align

    # Booking info
    info_rows = [
        ('Booking No.',         booking.booking_number),
        ('Customer',            booking.customer.name),
        ('Booking Date',        booking.booking_date.strftime('%d %b %Y') if booking.booking_date else ''),
        ('Order Sending Date',  booking.order_sending_date.strftime('%d %b %Y') if booking.order_sending_date else '—'),
        ('Generated At',        timezone.now().strftime('%d %b %Y %H:%M')),
    ]
    row = 3
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value).font = normal_font
        row += 1

    # Empty row
    row += 1

    # Items header
    headers = ['#', 'Product', 'Category', 'Qty', 'Unit Price (Rs.)', 'Disc%', 'Disc Amt (Rs.)', 'Line Total (Rs.)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = white_bold
        cell.fill      = blue_fill
        cell.alignment = right_align if col > 3 else Alignment(horizontal='left')
        cell.border    = thin_border
    row += 1

    # Items
    for idx, item in enumerate(booking.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=item.product.name).border = thin_border
        ws.cell(row=row, column=3, value=item.product.category.name).border = thin_border
        ws.cell(row=row, column=4, value=float(item.quantity)).alignment = right_align
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5, value=float(item.unit_price)).alignment = right_align
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6, value=float(item.discount_percent)).alignment = right_align
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=7, value=float(item.discount_amount)).alignment = right_align
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=8, value=float(item.line_total)).alignment = right_align
        ws.cell(row=row, column=8).border = thin_border
        row += 1

    row += 1  # empty row

    # Totals
    totals = [
        ('Subtotal (Rs.)',       float(booking.subtotal)),
        ('Total Discount (Rs.)', float(booking.discount_amount)),
        ('Grand Total (Rs.)',    float(booking.total_amount)),
    ]
    for label, value in totals:
        ws.cell(row=row, column=6, value=label).font = bold_font
        ws.cell(row=row, column=6).alignment = right_align
        val_cell = ws.cell(row=row, column=8, value=value)
        val_cell.font      = bold_font
        val_cell.alignment = right_align
        row += 1

    # Column widths
    col_widths = [5, 28, 18, 10, 18, 10, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"quotation_{booking.booking_number}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────── export Excel ────────────────────────────────────

def booking_export_excel(request):
    q         = request.GET.get('q', '').strip()
    customer  = request.GET.get('customer', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    status    = request.GET.get('status', '').strip()

    qs = BookingOrder.objects.select_related('customer', 'bill').order_by('-booking_date', '-id')
    qs = _apply_booking_filters(qs, q, customer, date_from, date_to, status)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Booking Orders'

    blue_fill  = PatternFill('solid', fgColor='1D4ED8')
    white_bold = Font(color='FFFFFF', bold=True, size=10)
    bold_font  = Font(bold=True, size=10)
    right_align= Alignment(horizontal='right')

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Booking Orders Report — SENOVKA ERP'
    ws['A1'].font = Font(bold=True, size=14, color='1D4ED8')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f"Generated: {timezone.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].font = Font(size=9, color='6B7280')

    headers = ['#', 'Booking No.', 'Customer', 'Booking Date', 'Order Sending Date', 'Status', 'Total (Rs.)', 'Bill No.']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = white_bold
        cell.fill = blue_fill
        cell.alignment = right_align if col == 7 else Alignment(horizontal='left')

    for row_idx, b in enumerate(qs, 1):
        row = row_idx + 4
        ws.cell(row=row, column=1, value=row_idx)
        ws.cell(row=row, column=2, value=b.booking_number)
        ws.cell(row=row, column=3, value=b.customer.name)
        ws.cell(row=row, column=4, value=b.booking_date.strftime('%d-%m-%Y') if b.booking_date else '')
        ws.cell(row=row, column=5, value=b.order_sending_date.strftime('%d-%m-%Y') if b.order_sending_date else '—')
        ws.cell(row=row, column=6, value=b.get_status_display())
        ws.cell(row=row, column=7, value=float(b.total_amount)).alignment = right_align
        ws.cell(row=row, column=8, value=b.bill.bill_number if b.bill else '—')

    col_widths = [5, 22, 28, 14, 18, 12, 15, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="booking_orders.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────── export PDF ──────────────────────────────────────

def booking_export_pdf(request):
    q         = request.GET.get('q', '').strip()
    customer  = request.GET.get('customer', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    status    = request.GET.get('status', '').strip()

    qs = BookingOrder.objects.select_related('customer', 'bill').order_by('-booking_date', '-id')
    qs = _apply_booking_filters(qs, q, customer, date_from, date_to, status)

    agg          = qs.aggregate(total=Sum('total_amount'))
    total_amount = agg['total'] or Decimal('0')
    total_records= qs.count()

    from xhtml2pdf import pisa

    template = get_template('booking/booking_report_pdf.html')
    html     = template.render({
        'bookings':      list(qs),
        'total_records': total_records,
        'total_amount':  total_amount,
        'generated_at':  timezone.now(),
        'q':         q,
        'customer':  customer,
        'date_from': date_from,
        'date_to':   date_to,
        'status':    status,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="booking_orders.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response


# ─────────────────────── production needs API ─────────────────────────────────

from django.views.decorators.http import require_POST as _require_POST
from django.http import JsonResponse as _JsonResponse


@_require_POST
def api_production_needs(request):
    try:
        payload = json.loads(request.body)
        items   = payload.get('items', [])
    except (json.JSONDecodeError, AttributeError):
        return _JsonResponse({'error': 'Invalid JSON'}, status=400)

    result = []
    for raw in items:
        try:
            product_id  = int(raw['product_id'])
            current_qty = Decimal(str(raw.get('quantity', 0)))
        except (KeyError, ValueError, InvalidOperation):
            continue

        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            continue

        pending_qty = (
            BookingItem.objects
            .filter(product_id=product_id, booking__status=BookingOrder.PENDING)
            .aggregate(total=Sum('quantity'))['total']
        ) or Decimal('0')

        total_ordered     = pending_qty + current_qty
        production_needed = max(Decimal('0'), total_ordered - product.qty)

        result.append({
            'product_id':       product_id,
            'product_name':     product.name,
            'available_qty':    float(product.qty),
            'pending_other_qty':float(pending_qty),
            'current_qty':      float(current_qty),
            'total_ordered':    float(total_ordered),
            'production_needed':float(production_needed),
        })

    return _JsonResponse({'items': result})
