from decimal import Decimal

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone

from .forms import MaterialPurchaseForm
from .models import MaterialPurchase
from erp.utils import log_activity, current_month_bounds


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_filters(request):
    q            = request.GET.get('q', '').strip()
    supplier     = request.GET.get('supplier', '').strip()
    date_from    = request.GET.get('date_from', '').strip()
    date_to      = request.GET.get('date_to', '').strip()
    month        = request.GET.get('month', '').strip()
    year         = request.GET.get('year', '').strip()
    unit_type    = request.GET.get('unit_type', '').strip()

    # Default to the current month only on a fresh visit (none of these
    # date-ish params present at all) -- never alongside an explicit
    # month/year, which would otherwise combine into an impossible range.
    no_date_param = not any(k in request.GET for k in ('date_from', 'date_to', 'month', 'year'))
    if no_date_param:
        date_from, date_to = current_month_bounds()

    return q, supplier, date_from, date_to, month, year, unit_type


def _apply_filters(qs, q, supplier, date_from, date_to, month, year, unit_type):
    if q:
        qs = qs.filter(
            Q(invoice_number__icontains=q) |
            Q(supplier_name__icontains=q)  |
            Q(material_name__icontains=q)
        )
    if supplier:
        qs = qs.filter(supplier_name__icontains=supplier)
    if date_from:
        qs = qs.filter(received_date__gte=date_from)
    if date_to:
        qs = qs.filter(received_date__lte=date_to)
    if month and year:
        try:
            qs = qs.filter(received_date__month=int(month), received_date__year=int(year))
        except ValueError:
            pass
    elif month:
        try:
            qs = qs.filter(received_date__month=int(month))
        except ValueError:
            pass
    elif year:
        try:
            qs = qs.filter(received_date__year=int(year))
        except ValueError:
            pass
    if unit_type in (MaterialPurchase.KG, MaterialPurchase.G):
        qs = qs.filter(unit_type=unit_type)
    return qs


def _filtered_qs(request):
    filters = _parse_filters(request)
    qs = MaterialPurchase.objects.all()
    return _apply_filters(qs, *filters), filters


def _distinct_suppliers():
    return (
        MaterialPurchase.objects.values_list('supplier_name', flat=True)
        .distinct().order_by('supplier_name')
    )


# ── List ──────────────────────────────────────────────────────────────────────

def purchase_list(request):
    qs, (q, supplier, date_from, date_to, month, year, unit_type) = _filtered_qs(request)

    total_amount   = qs.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    total_quantity = qs.aggregate(s=Sum('quantity'))['s'] or Decimal('0')
    total_records  = qs.count()

    paginator  = Paginator(qs, 15)
    page_obj   = paginator.get_page(request.GET.get('page'))

    today = timezone.localdate()
    month_choices = [
        (1,'January'),(2,'February'),(3,'March'),(4,'April'),
        (5,'May'),(6,'June'),(7,'July'),(8,'August'),
        (9,'September'),(10,'October'),(11,'November'),(12,'December'),
    ]
    year_choices = list(range(today.year, today.year - 6, -1))

    return render(request, 'purchasing/purchase_list.html', {
        'page_obj': page_obj,
        'purchases': page_obj.object_list,
        'total_amount': total_amount,
        'total_quantity': total_quantity,
        'total_records': total_records,
        'suppliers': _distinct_suppliers(),
        # filters
        'q': q,
        'supplier': supplier,
        'date_from': date_from,
        'date_to': date_to,
        'month': month,
        'year': year,
        'unit_type': unit_type,
        'month_choices': month_choices,
        'year_choices': year_choices,
    })


# ── Create ────────────────────────────────────────────────────────────────────

def purchase_create(request):
    if request.method == 'POST':
        form = MaterialPurchaseForm(request.POST)
        if form.is_valid():
            purchase = form.save()
            from django.urls import reverse
            log_activity(
                request, 'purchasing', 'purchase_created',
                f"Material purchase recorded: {purchase.material_name} | Supplier: {purchase.supplier_name} | Invoice: {purchase.invoice_number} | Rs. {purchase.total_amount}",
                reverse('purchasing:purchase_list'),
                related_id=purchase.pk,
            )
            messages.success(request, f'Purchase "{purchase.invoice_number}" recorded successfully.')
            return redirect('purchasing:purchase_list')
        for field, errs in form.errors.items():
            for err in errs:
                label = form.fields[field].label or field if field != '__all__' else ''
                messages.error(request, f'{label}: {err}' if label else err)
    else:
        form = MaterialPurchaseForm()

    return render(request, 'purchasing/purchase_form.html', {
        'form': form,
        'is_create': True,
        'title': 'Record New Purchase',
    })


# ── Update ────────────────────────────────────────────────────────────────────

def purchase_update(request, pk):
    purchase = get_object_or_404(MaterialPurchase, pk=pk)
    if request.method == 'POST':
        form = MaterialPurchaseForm(request.POST, instance=purchase)
        if form.is_valid():
            purchase = form.save()
            messages.success(request, f'Purchase "{purchase.invoice_number}" updated successfully.')
            return redirect('purchasing:purchase_list')
        for field, errs in form.errors.items():
            for err in errs:
                label = form.fields[field].label or field if field != '__all__' else ''
                messages.error(request, f'{label}: {err}' if label else err)
    else:
        form = MaterialPurchaseForm(instance=purchase)

    return render(request, 'purchasing/purchase_form.html', {
        'form': form,
        'purchase': purchase,
        'is_create': False,
        'title': f'Edit Purchase — {purchase.invoice_number}',
    })


# ── Delete ────────────────────────────────────────────────────────────────────

def purchase_delete(request, pk):
    purchase = get_object_or_404(MaterialPurchase, pk=pk)
    if request.method == 'POST':
        inv = purchase.invoice_number
        name = purchase.material_name
        purchase.delete()
        from django.urls import reverse
        log_activity(
            request, 'purchasing', 'purchase_deleted',
            f"Material purchase deleted: {name} | Invoice: {inv}",
            reverse('purchasing:purchase_list'),
        )
        messages.success(request, f'Purchase "{inv}" deleted.')
        return redirect('purchasing:purchase_list')
    return render(request, 'purchasing/purchase_confirm_delete.html', {'purchase': purchase})


# ── Excel Export ──────────────────────────────────────────────────────────────

def purchase_export_excel(request):
    qs, _ = _filtered_qs(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl is required for Excel export.', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Material Purchases'

    generated_at = timezone.localtime()
    total = qs.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')

    # Metadata
    ws.append(['Material Purchasing Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', generated_at.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Records', qs.count()])
    ws.append(['Total Amount (Rs.)', float(total)])
    ws.append([])

    # Header row
    headers = [
        '#', 'Invoice No.', 'Supplier', 'Material', 'Received Date',
        'Qty', 'Unit', 'Unit Price (Rs.)', 'Total (Rs.)',
        'Scale Weight', 'Discrepancy', 'Verified', 'Notes',
    ]
    ws.append(headers)
    header_row = ws.max_row
    header_fill = PatternFill('solid', fgColor='2563EB')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for i, p in enumerate(qs, 1):
        disc = p.weight_discrepancy
        ws.append([
            i,
            p.invoice_number,
            p.supplier_name,
            p.material_name,
            p.received_date.strftime('%d-%m-%Y'),
            float(p.quantity),
            p.unit_type,
            float(p.unit_price),
            float(p.total_amount),
            float(p.scale_weight) if p.scale_weight is not None else '',
            float(disc) if disc is not None else '',
            'Yes' if p.weight_verified else 'No',
            p.notes or '',
        ])

    # Column widths
    for col, width in zip('ABCDEFGHIJKLM', [4, 16, 22, 22, 14, 10, 6, 14, 14, 12, 12, 9, 30]):
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = generated_at.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="material_purchases_{ts}.xlsx"'
    wb.save(response)
    return response


# ── PDF Export ────────────────────────────────────────────────────────────────

def purchase_export_pdf(request):
    qs, (q, supplier, date_from, date_to, month, year, unit_type) = _filtered_qs(request)
    total = qs.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')

    context = {
        'purchases': qs,
        'total_amount': total,
        'total_records': qs.count(),
        'generated_at': timezone.localtime(),
        'q': q,
        'supplier': supplier,
        'date_from': date_from,
        'date_to': date_to,
        'month': month,
        'year': year,
    }

    template = get_template('purchasing/purchase_report_pdf.html')
    html = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="material_purchases_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF generation error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'purchasing/purchase_report_pdf.html', context)
