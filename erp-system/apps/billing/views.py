import json
from datetime import timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Exists, F, OuterRef, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.urls import reverse
from django.utils import timezone

from apps.customers.models import Customer, CustomerLedger, CustomerProductPrice, add_ledger_entry
from apps.production.models import Product
from .models import Bill, BillItem, HeldBill, Payment
from erp.decorators import superadmin_required
from erp.models import CompanySettings
from erp.utils import log_activity, resolve_date_range


# ─────────────────────────── helpers ──────────────────────────────────────────

_TWO = Decimal('0.01')


def _fmt_qty(q):
    return f'{q:.2f}'.rstrip('0').rstrip('.')




def _d(value, default='0'):
    try:
        return Decimal(str(value or default)).quantize(_TWO, rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError):
        return Decimal(default)


def _parse_filters(request):
    date_from, date_to = resolve_date_range(request)
    return {
        'q':          request.GET.get('q', '').strip(),
        'status':     request.GET.get('status', '').strip(),
        'customer':   request.GET.get('customer', '').strip(),
        'date_from':  date_from,
        'date_to':    date_to,
        'method':     request.GET.get('method', '').strip(),
    }


def _apply_filters(qs, f):
    if f['q']:
        qs = qs.filter(
            Q(bill_number__icontains=f['q']) | Q(customer__name__icontains=f['q'])
        )
    if f['status']:
        qs = qs.filter(status=f['status'])
    if f['customer']:
        qs = qs.filter(customer__name__icontains=f['customer'])
    if f['date_from']:
        qs = qs.filter(bill_date__gte=f['date_from'])
    if f['date_to']:
        qs = qs.filter(bill_date__lte=f['date_to'])
    if f['method']:
        qs = qs.filter(payment_method=f['method'])
    return qs


# ─────────────────────────── AJAX ─────────────────────────────────────────────

def api_customer_info(request, pk):
    """Return customer balance + all products for the billing form."""
    customer = get_object_or_404(Customer, pk=pk)

    price_map = {
        cpp.product_id: float(cpp.unit_price)
        for cpp in CustomerProductPrice.objects.filter(customer=customer)
    }

    products = [
        {
            'id':              p.id,
            'name':            p.name,
            'category':        p.category.name,
            'size':            p.size or '',
            'unit_price':      price_map.get(p.id, 0),
            'has_custom_price': p.id in price_map,
            'available_qty':   float(p.qty),
        }
        for p in Product.objects.select_related('category').order_by('name')
    ]

    return JsonResponse({
        'balance':  float(customer.balance),
        'products': products,
    })


# ─────────────────────────── bill list ────────────────────────────────────────

def bill_list(request):
    f   = _parse_filters(request)
    qs  = _apply_filters(
        Bill.objects.select_related('customer').all(), f
    )

    agg          = qs.aggregate(total=Sum('total_amount'), due=Sum('amount_due'))
    total_amount = agg['total'] or Decimal('0')
    total_due    = agg['due']   or Decimal('0')
    total_records= qs.count()

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    customers = Customer.objects.order_by('name')
    today     = timezone.localdate()

    return render(request, 'billing/bill_list.html', {
        'page_obj':      page_obj,
        'bills':         page_obj.object_list,
        'total_amount':  total_amount,
        'total_due':     total_due,
        'total_records': total_records,
        'customers':     customers,
        'method_choices': Bill.METHOD_CHOICES,
        **f,
    })


# ─────────────────────────── create bill ──────────────────────────────────────

def bill_create(request):
    customers = Customer.objects.order_by('name')

    if request.method == 'POST':
        try:
            bill = _create_bill(request)
            log_activity(
                request, 'billing', 'bill_created',
                f"Bill created: {bill.bill_number} | {bill.customer.name} | Rs. {bill.total_amount}",
                reverse('billing:bill_detail', kwargs={'pk': bill.pk}),
                related_id=bill.pk,
            )
            messages.success(request, f'Bill {bill.bill_number} created successfully.')
            return redirect('billing:bill_detail', pk=bill.pk)
        except (ValueError, Product.DoesNotExist, Customer.DoesNotExist) as exc:
            messages.error(request, str(exc))

    recall_data = request.session.pop('recall_bill_data', None)

    return render(request, 'billing/bill_create.html', {
        'customers': customers,
        'today':     timezone.localdate().isoformat(),
        'method_choices': Bill.METHOD_CHOICES,
        'recall_data': recall_data,
        'held_bills_count': HeldBill.objects.count(),
    })


@transaction.atomic
def _create_bill(request):
    POST = request.POST

    # ── customer ──────────────────────────────────────────────
    customer_id = POST.get('customer_id', '').strip()
    if not customer_id:
        raise ValueError('Please select a customer.')
    customer = Customer.objects.select_for_update().get(pk=customer_id)

    # ── date ─────────────────────────────────────────────────
    from datetime import datetime
    date_str  = POST.get('bill_date', '') or timezone.localdate().isoformat()
    try:
        bill_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        bill_date = timezone.localdate()

    # ── payment method ────────────────────────────────────────
    method = POST.get('payment_method', '')
    if method not in dict(Bill.METHOD_CHOICES):
        raise ValueError('Please select a valid payment method.')

    # ── items ─────────────────────────────────────────────────
    try:
        items_raw = json.loads(POST.get('items_json', '[]') or '[]')
    except json.JSONDecodeError:
        raise ValueError('Invalid items data — please refresh and try again.')
    if not items_raw:
        raise ValueError('Please add at least one product to the bill.')

    subtotal       = Decimal('0')
    total_discount = Decimal('0')
    validated      = []

    for raw in items_raw:
        pid = raw.get('product_id')
        product = Product.objects.select_for_update().get(pk=pid)

        qty        = _d(raw.get('quantity'), '0')
        unit_price = _d(raw.get('unit_price'), '0')
        disc_pct   = _d(raw.get('discount_percent'), '0')

        if qty <= 0:
            raise ValueError(f'Quantity for "{product.name}" must be > 0.')
        if unit_price < 0:
            raise ValueError(f'Unit price for "{product.name}" cannot be negative.')
        if not (0 <= disc_pct <= 100):
            raise ValueError(f'Discount for "{product.name}" must be 0–100%.')
        if product.qty < qty:
            raise ValueError(
                f'Insufficient stock for "{product.name}". '
                f'Available: {_fmt_qty(product.qty)}, Requested: {_fmt_qty(qty)}.'
            )

        disc_amt   = (qty * unit_price * disc_pct / 100).quantize(_TWO, rounding=ROUND_HALF_UP)
        line_total = (qty * unit_price - disc_amt).quantize(_TWO, rounding=ROUND_HALF_UP)

        subtotal       += (qty * unit_price).quantize(_TWO, rounding=ROUND_HALF_UP)
        total_discount += disc_amt
        validated.append({
            'product':          product,
            'quantity':         qty,
            'unit_price':       unit_price,
            'discount_percent': disc_pct,
            'discount_amount':  disc_amt,
            'line_total':       line_total,
        })

    item_total    = (subtotal - total_discount).quantize(_TWO, rounding=ROUND_HALF_UP)
    bill_disc_pct = _d(POST.get('bill_discount_percent', '0'))
    if not (Decimal('0') <= bill_disc_pct <= Decimal('100')):
        raise ValueError('Bill discount must be between 0 and 100%.')
    bill_disc_amt = (item_total * bill_disc_pct / 100).quantize(_TWO, rounding=ROUND_HALF_UP)
    total_amount  = (item_total - bill_disc_amt).quantize(_TWO, rounding=ROUND_HALF_UP)
    credit        = max(Decimal('0'), customer.balance)

    # ── resolve payment ───────────────────────────────────────
    status      = Bill.COMPLETED
    balance_used= Decimal('0')
    amount_paid = Decimal('0')
    amount_due  = Decimal('0')
    overpayment = Decimal('0')
    pay_records = []   # list[dict] to create after bill save

    if method == Bill.PAY_LATER:
        # Check outstanding limit before allowing this bill
        if customer.max_outstanding_limit is not None:
            projected_balance = customer.balance - total_amount
            if projected_balance < -customer.max_outstanding_limit:
                raise ValueError(
                    f"{customer.name} has a maximum outstanding limit of "
                    f"Rs. {customer.max_outstanding_limit:,.2f}. "
                    f"Current balance: Rs. {customer.balance:,.2f}, "
                    f"this bill: Rs. {total_amount:,.2f}. "
                    f"Bill blocked — reduce outstanding first."
                )
        status     = Bill.PENDING
        amount_due = total_amount

    elif method in (Bill.FULL_CASH, Bill.FULL_CHEQUE, Bill.MIXED):
        balance_used    = min(credit, total_amount)
        need_to_collect = (total_amount - balance_used).quantize(_TWO, rounding=ROUND_HALF_UP)

        cash_in        = Decimal('0')
        cheque_in      = Decimal('0')
        cheque_records = []

        if method in (Bill.FULL_CASH, Bill.MIXED):
            cash_in = _d(POST.get('cash_amount'))
        if method in (Bill.FULL_CHEQUE, Bill.MIXED):
            cheque_records, cheque_in = _parse_cheques(POST, customer.name)

        collected = (cash_in + cheque_in).quantize(_TWO, rounding=ROUND_HALF_UP)

        if need_to_collect > 0 and collected < need_to_collect:
            raise ValueError(
                f'Underpayment not allowed in full-payment mode. '
                f'Need Rs. {need_to_collect:,.2f}, received Rs. {collected:,.2f}.'
            )

        overpayment = max(Decimal('0'), collected - need_to_collect)
        amount_paid = collected

        senovka = POST.get('cash_is_senovka_transfer') == '1'
        if balance_used > 0:
            pay_records.append({'method': Payment.BALANCE, 'amount': balance_used})
        if cash_in > 0:
            pay_records.append({'method': Payment.CASH, 'amount': cash_in, 'is_senovka_transfer': senovka})
        if cheque_records:
            pay_records.extend(cheque_records)

    elif method in (Bill.PARTIAL_CASH, Bill.PARTIAL_CHEQUE):
        cash_in        = _d(POST.get('cash_amount')) if method == Bill.PARTIAL_CASH else Decimal('0')
        cheque_records = []
        cheque_in      = Decimal('0')
        if method == Bill.PARTIAL_CHEQUE:
            cheque_records, cheque_in = _parse_cheques(POST, customer.name)
        collected = cash_in + cheque_in

        if collected <= 0:
            raise ValueError('Please enter the partial payment amount.')
        if collected >= total_amount:
            # treated as full payment
            overpayment = (collected - total_amount).quantize(_TWO, rounding=ROUND_HALF_UP)
            amount_paid = total_amount
            amount_due  = Decimal('0')
        else:
            amount_paid = collected
            amount_due  = (total_amount - collected).quantize(_TWO, rounding=ROUND_HALF_UP)
            status      = Bill.PENDING

        senovka = POST.get('cash_is_senovka_transfer') == '1'
        if cash_in > 0:
            pay_records.append({'method': Payment.CASH, 'amount': cash_in, 'is_senovka_transfer': senovka})
        if cheque_records:
            pay_records.extend(cheque_records)

    # ── persist bill ──────────────────────────────────────────
    bill = Bill.objects.create(
        customer              = customer,
        bill_date             = bill_date,
        status                = status,
        payment_method        = method,
        subtotal              = subtotal,
        discount_amount       = total_discount,
        bill_discount_percent = bill_disc_pct,
        bill_discount_amount  = bill_disc_amt,
        total_amount          = total_amount,
        balance_used          = balance_used,
        amount_paid           = amount_paid,
        amount_due            = amount_due,
        notes                 = POST.get('notes', '').strip(),
        created_by            = request.user if request.user.is_authenticated else None,
    )

    # ── bill items + inventory ────────────────────────────────
    for raw, item in zip(items_raw, validated):
        BillItem.objects.create(
            bill            = bill,
            product         = item['product'],
            quantity        = item['quantity'],
            unit_price      = item['unit_price'],
            discount_percent= item['discount_percent'],
            discount_amount = item['discount_amount'],
            line_total      = item['line_total'],
        )
        Product.objects.filter(pk=item['product'].pk).update(qty=F('qty') - item['quantity'])

        # Save/update the customer-specific price if requested and price > 0.
        # update_or_create (not get_or_create) so editing an already-saved price overwrites it.
        if raw.get('save_price') and item['unit_price'] > 0:
            CustomerProductPrice.objects.update_or_create(
                customer=customer,
                product=item['product'],
                defaults={'unit_price': item['unit_price']},
            )

    # ── payment records ───────────────────────────────────────
    for pd in pay_records:
        _save_payment(bill, bill_date, pd)

    # ── customer balance ──────────────────────────────────────
    if method == Bill.PAY_LATER:
        customer.balance = customer.balance - total_amount
    else:
        customer.balance = customer.balance - balance_used + overpayment
    customer.save(update_fields=['balance'])

    # ── ledger entry ──────────────────────────────────────────
    _bill_create_ledger(
        customer, bill, bill_date, method, total_amount,
        balance_used, overpayment,
        cash_in if method != Bill.PAY_LATER else Decimal('0'),
        cheque_in if method != Bill.PAY_LATER else Decimal('0'),
    )
    return bill


def _bill_create_ledger(customer, bill, bill_date, method, total_amount,
                        balance_used, overpayment, cash_in, cheque_in):
    _D = Decimal('0')

    if method == Bill.PAY_LATER:
        add_ledger_entry(
            customer, date=bill_date, bill=bill,
            description=f"Sale on Credit - {bill.bill_number}",
            transaction_type=CustomerLedger.SALE,
            debit=total_amount, credit=_D,
        )
    elif method in (Bill.PARTIAL_CASH, Bill.PARTIAL_CHEQUE):
        collected = cash_in + cheque_in
        label = 'Cash' if method == Bill.PARTIAL_CASH else 'Cheque'
        add_ledger_entry(
            customer, date=bill_date, bill=bill,
            description=f"Sale ({label}, Partial) - {bill.bill_number}",
            transaction_type=CustomerLedger.SALE_PARTIAL,
            debit=total_amount, credit=collected,
        )
    else:
        # FULL_CASH / FULL_CHEQUE / MIXED – net balance change = -balance_used + overpayment
        labels = {Bill.FULL_CASH: 'Cash', Bill.FULL_CHEQUE: 'Cheque', Bill.MIXED: 'Mixed'}
        label = labels.get(method, 'Cash')

        if balance_used > _D and overpayment > _D:
            add_ledger_entry(
                customer, date=bill_date, bill=bill,
                description=f"Sale ({label}) - {bill.bill_number}",
                transaction_type=CustomerLedger.SALE_CASH,
                debit=total_amount + balance_used, credit=total_amount + overpayment,
            )
        elif balance_used > _D:
            add_ledger_entry(
                customer, date=bill_date, bill=bill,
                description=f"Sale ({label}, Credit Applied Rs.{balance_used:,.2f}) - {bill.bill_number}",
                transaction_type=CustomerLedger.SALE_CASH,
                debit=total_amount + balance_used, credit=total_amount,
            )
        elif overpayment > _D:
            add_ledger_entry(
                customer, date=bill_date, bill=bill,
                description=f"Sale ({label}, Overpayment Rs.{overpayment:,.2f}) - {bill.bill_number}",
                transaction_type=CustomerLedger.SALE_CASH,
                debit=total_amount, credit=total_amount + overpayment,
            )
        else:
            add_ledger_entry(
                customer, date=bill_date, bill=bill,
                description=f"Sale ({label}) - {bill.bill_number}",
                transaction_type=CustomerLedger.SALE_CASH,
                debit=total_amount, credit=total_amount,
            )


def _parse_cheques(POST, default_customer_name):
    """Parse the 'cheques_json' field (list of cheque dicts) into Payment.CHEQUE
    pay_records, one per cheque, so a single payment can be split across several
    cheques. Returns (records, total_amount)."""
    from datetime import datetime

    def _dt(val):
        try:
            return datetime.strptime(val, '%Y-%m-%d').date() if val else None
        except (ValueError, TypeError):
            return None

    try:
        rows = json.loads(POST.get('cheques_json', '[]') or '[]')
    except json.JSONDecodeError:
        raise ValueError('Invalid cheque data — please refresh and try again.')

    records = []
    total   = Decimal('0')
    for row in rows:
        amount = _d(row.get('amount'), '0')
        if amount <= 0:
            continue
        cheque_number = (row.get('cheque_number') or '').strip()
        bank_name     = (row.get('bank_name') or '').strip()
        if not cheque_number:
            raise ValueError('Cheque number is required for every cheque.')
        if not bank_name:
            raise ValueError('Bank name is required for every cheque.')

        records.append({
            'method':         Payment.CHEQUE,
            'amount':         amount,
            'customer_name':  (row.get('customer_name') or '').strip() or default_customer_name,
            'cheque_number':  cheque_number,
            'bank_name':      bank_name,
            'branch_name':    (row.get('branch_name') or '').strip(),
            'account_number': (row.get('account_number') or '').strip(),
            'received_date':  _dt(row.get('received_date')),
            'maturity_date':  _dt(row.get('maturity_date')),
            'cheque_status':  Payment.CHQ_PENDING,
        })
        total += amount

    return records, total


def _save_payment(bill, payment_date, pd):
    p = Payment(
        bill        = bill,
        payment_date= payment_date,
        method      = pd['method'],
        amount      = pd['amount'],
    )
    for field in (
        'customer_name', 'cheque_number', 'bank_name', 'branch_name',
        'account_number', 'received_date', 'maturity_date', 'cheque_status',
        'is_senovka_transfer',
    ):
        if field in pd:
            setattr(p, field, pd[field])
    p.save()


# ─────────────────────────── bill detail ──────────────────────────────────────

def bill_detail(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related('customer', 'created_by')
            .prefetch_related('items__product__category', 'payments'),
        pk=pk,
    )
    is_superadmin = False
    try:
        is_superadmin = request.user.profile.is_superadmin()
    except Exception:
        pass
    return render(request, 'billing/bill_detail.html', {'bill': bill, 'is_superadmin': is_superadmin})


# ─────────────────────────── cancel / delete ──────────────────────────────────

def bill_cancel(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    if request.method != 'POST':
        return render(request, 'billing/bill_confirm_cancel.html', {'bill': bill})

    if bill.status == Bill.COMPLETED:
        messages.error(request, 'Completed bills cannot be cancelled.')
        return redirect('billing:bill_detail', pk=pk)

    with transaction.atomic():
        # Restore inventory
        for item in bill.items.select_related('product'):
            Product.objects.filter(pk=item.product.pk).update(qty=F('qty') + item.quantity)

        customer = Customer.objects.select_for_update().get(pk=bill.customer.pk)
        if bill.payment_method == Bill.PAY_LATER:
            # Reverse the net balance impact of the bill and any partial settlements.
            # Creation decremented by total_amount; settlements incremented by (collected - balance_now).
            # Net effect = total_amount - amount_paid + balance_used; restore by adding it back.
            customer.balance = customer.balance + bill.total_amount - bill.amount_paid + bill.balance_used
        else:
            overpay = max(Decimal('0'), bill.amount_paid - (bill.total_amount - bill.balance_used))
            customer.balance = customer.balance + bill.balance_used - overpay
        customer.save(update_fields=['balance'])

        bill.status = Bill.CANCELLED
        bill.save(update_fields=['status'])

    log_activity(
        request, 'billing', 'bill_cancelled',
        f"Bill cancelled: {bill.bill_number} | {bill.customer.name} | Rs. {bill.total_amount}",
        reverse('billing:bill_detail', kwargs={'pk': bill.pk}),
    )
    messages.success(request, f'Bill {bill.bill_number} cancelled and inventory restored.')
    return redirect('billing:bill_list')


# ─────────────────────────── edit bill ────────────────────────────────────────

@superadmin_required
def bill_edit(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related('customer')
            .prefetch_related('items__product__category'),
        pk=pk,
    )

    if bill.status == Bill.CANCELLED:
        messages.error(request, 'Cancelled bills cannot be edited.')
        return redirect('billing:bill_detail', pk=pk)

    if request.method == 'POST':
        try:
            _apply_bill_edit(request, bill)
            log_activity(
                request, 'billing', 'bill_edited',
                f"Bill edited: {bill.bill_number} | {bill.customer.name}",
                reverse('billing:bill_detail', kwargs={'pk': bill.pk}),
                related_id=bill.pk,
            )
            messages.success(request, f'Bill {bill.bill_number} updated successfully.')
            return redirect('billing:bill_detail', pk=pk)
        except (ValueError, Product.DoesNotExist) as exc:
            messages.error(request, str(exc))

    # Build product data with customer-specific prices
    price_map = {
        cpp.product_id: float(cpp.unit_price)
        for cpp in CustomerProductPrice.objects.filter(customer=bill.customer)
    }
    products_data = [
        {
            'id':              p.id,
            'name':            p.name,
            'category':        p.category.name,
            'size':            p.size or '',
            'unit_price':      price_map.get(p.id, 0),
            'has_custom_price': p.id in price_map,
            'available_qty':   float(p.qty),
        }
        for p in Product.objects.select_related('category').order_by('name')
    ]

    # Pre-populate existing items (restore available_qty to include current bill quantities)
    existing_items = []
    for item in bill.items.all():
        pid = item.product.id
        # The product's current qty does NOT include the already-deducted bill qty,
        # so show available + current item qty as the editable ceiling
        avail = float(Product.objects.values_list('qty', flat=True).get(pk=pid)) + float(item.quantity)
        existing_items.append({
            'product_id':       pid,
            'name':             item.product.name,
            'category':         item.product.category.name,
            'available_qty':    avail,
            'unit_price':       float(item.unit_price),
            'has_custom_price': pid in price_map,
            'save_price':       False,
            'quantity':         float(item.quantity),
            'discount_percent': float(item.discount_percent),
            'discount_amount':  float(item.discount_amount),
            'line_total':       float(item.line_total),
        })

    return render(request, 'billing/bill_edit.html', {
        'bill':           bill,
        'products_json':  json.dumps(products_data),
        'existing_items': json.dumps(existing_items),
        'today':          timezone.localdate().isoformat(),
    })


@transaction.atomic
def _apply_bill_edit(request, bill):
    from datetime import datetime
    POST = request.POST

    bill = Bill.objects.select_for_update().get(pk=bill.pk)

    try:
        items_raw = json.loads(POST.get('items_json', '[]') or '[]')
    except json.JSONDecodeError:
        raise ValueError('Invalid items data — please refresh and try again.')
    if not items_raw:
        raise ValueError('Please add at least one product to the bill.')

    # ── Restore old inventory ─────────────────────────────────
    for old_item in bill.items.select_related('product'):
        Product.objects.filter(pk=old_item.product.pk).update(qty=F('qty') + old_item.quantity)

    # ── Validate new items ────────────────────────────────────
    subtotal       = Decimal('0')
    total_discount = Decimal('0')
    validated      = []

    for raw in items_raw:
        pid     = raw.get('product_id')
        product = Product.objects.select_for_update().get(pk=pid)
        qty        = _d(raw.get('quantity'), '0')
        unit_price = _d(raw.get('unit_price'), '0')
        disc_pct   = _d(raw.get('discount_percent'), '0')

        if qty <= 0:
            raise ValueError(f'Quantity for "{product.name}" must be > 0.')
        if unit_price < 0:
            raise ValueError(f'Unit price for "{product.name}" cannot be negative.')
        if not (0 <= disc_pct <= 100):
            raise ValueError(f'Discount for "{product.name}" must be 0–100%.')
        if product.qty < qty:
            raise ValueError(
                f'Insufficient stock for "{product.name}". '
                f'Available: {_fmt_qty(product.qty)}, Requested: {_fmt_qty(qty)}.'
            )

        disc_amt   = (qty * unit_price * disc_pct / 100).quantize(_TWO, rounding=ROUND_HALF_UP)
        line_total = (qty * unit_price - disc_amt).quantize(_TWO, rounding=ROUND_HALF_UP)

        subtotal       += (qty * unit_price).quantize(_TWO, rounding=ROUND_HALF_UP)
        total_discount += disc_amt
        validated.append({
            'product':          product,
            'quantity':         qty,
            'unit_price':       unit_price,
            'discount_percent': disc_pct,
            'discount_amount':  disc_amt,
            'line_total':       line_total,
        })

    new_total = (subtotal - total_discount).quantize(_TWO, rounding=ROUND_HALF_UP)

    # ── Replace items ─────────────────────────────────────────
    bill.items.all().delete()
    for raw, item in zip(items_raw, validated):
        BillItem.objects.create(
            bill            = bill,
            product         = item['product'],
            quantity        = item['quantity'],
            unit_price      = item['unit_price'],
            discount_percent= item['discount_percent'],
            discount_amount = item['discount_amount'],
            line_total      = item['line_total'],
        )
        Product.objects.filter(pk=item['product'].pk).update(qty=F('qty') - item['quantity'])

        if raw.get('save_price') and item['unit_price'] > 0:
            CustomerProductPrice.objects.update_or_create(
                customer=bill.customer,
                product=item['product'],
                defaults={'unit_price': item['unit_price']},
            )

    # ── Recalculate totals ────────────────────────────────────
    amount_collected = bill.amount_paid + bill.balance_used
    new_amount_due   = max(Decimal('0'), new_total - amount_collected).quantize(_TWO, rounding=ROUND_HALF_UP)
    new_status       = Bill.COMPLETED if new_amount_due == Decimal('0') else Bill.PENDING

    # ── Update date / notes ───────────────────────────────────
    date_str  = POST.get('bill_date', '').strip() or bill.bill_date.isoformat()
    try:
        new_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        new_date = bill.bill_date

    bill.subtotal        = subtotal
    bill.discount_amount = total_discount
    bill.total_amount    = new_total
    bill.amount_due      = new_amount_due
    bill.status          = new_status
    bill.bill_date       = new_date
    bill.notes           = POST.get('notes', bill.notes).strip()
    bill.save(update_fields=[
        'subtotal', 'discount_amount', 'total_amount',
        'amount_due', 'status', 'bill_date', 'notes', 'updated_at',
    ])


# ─────────────────────────── hold / recall bills ───────────────────────────────

def hold_bill_create(request):
    """AJAX endpoint: park the in-progress bill-create cart as a HeldBill."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Invalid request.'}, status=405)

    POST = request.POST
    try:
        items_raw = json.loads(POST.get('items_json', '[]') or '[]')
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid items data — please refresh and try again.'}, status=400)
    if not items_raw:
        return JsonResponse({'ok': False, 'error': 'Add at least one product before holding.'}, status=400)

    customer_id = POST.get('customer_id', '').strip()
    customer = Customer.objects.filter(pk=customer_id).first() if customer_id else None
    customer_name = customer.name if customer else POST.get('customer_name', '').strip()

    from datetime import datetime
    date_str = POST.get('bill_date', '') or timezone.localdate().isoformat()
    try:
        bill_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        bill_date = timezone.localdate()

    total_amount = Decimal('0')
    for raw in items_raw:
        total_amount += _d(raw.get('line_total'), '0')

    held = HeldBill.objects.create(
        customer               = customer,
        customer_name          = customer_name,
        bill_date               = bill_date,
        notes                   = POST.get('notes', '').strip(),
        bill_discount_percent   = _d(POST.get('bill_discount_percent', '0')),
        items_json              = json.dumps(items_raw),
        item_count              = len(items_raw),
        total_amount            = total_amount,
        held_by                 = request.user if request.user.is_authenticated else None,
    )
    log_activity(
        request, 'billing', 'bill_held',
        f"Bill held: {held.customer_name or 'Walk-in'} | {held.item_count} item(s) | Rs. {held.total_amount}",
        reverse('billing:held_bills'),
        related_id=held.pk,
    )
    return JsonResponse({'ok': True, 'id': held.pk, 'held_bills_count': HeldBill.objects.count()})


def held_bills_list(request):
    holds = HeldBill.objects.select_related('customer', 'held_by')
    return render(request, 'billing/held_bills.html', {'holds': holds})


def held_bill_recall(request, pk):
    if request.method != 'POST':
        return redirect('billing:held_bills')

    held = get_object_or_404(HeldBill, pk=pk)
    try:
        items = json.loads(held.items_json or '[]')
    except json.JSONDecodeError:
        items = []

    request.session['recall_bill_data'] = {
        'customer_id':            held.customer_id,
        'customer_name':          held.customer_name,
        'bill_date':              held.bill_date.isoformat(),
        'notes':                  held.notes,
        'bill_discount_percent':  str(held.bill_discount_percent),
        'items':                  items,
    }
    held.delete()
    return redirect('billing:bill_create')


def held_bill_delete(request, pk):
    held = get_object_or_404(HeldBill, pk=pk)
    if request.method == 'POST':
        label = held.customer_name or 'Walk-in'
        held.delete()
        log_activity(
            request, 'billing', 'bill_held_discarded',
            f"Held bill discarded: {label}",
            reverse('billing:held_bills'),
        )
        messages.success(request, f'Held bill for {label} discarded.')
    return redirect('billing:held_bills')


# ─────────────────────────── pending bills ────────────────────────────────────

def pending_bills(request):
    q           = request.GET.get('q', '').strip()
    customer_f  = request.GET.get('customer', '').strip()
    date_from   = request.GET.get('date_from', '').strip()
    date_to     = request.GET.get('date_to', '').strip()

    qs = Bill.objects.filter(status=Bill.PENDING).select_related('customer')
    if q:
        qs = qs.filter(
            Q(bill_number__icontains=q) | Q(customer__name__icontains=q)
        )
    if customer_f:
        qs = qs.filter(customer__name__icontains=customer_f)
    if date_from:
        qs = qs.filter(bill_date__gte=date_from)
    if date_to:
        qs = qs.filter(bill_date__lte=date_to)

    total_due    = qs.aggregate(s=Sum('amount_due'))['s'] or Decimal('0')
    total_records= qs.count()

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    customers = Customer.objects.filter(bills__status=Bill.PENDING).distinct().order_by('name')

    return render(request, 'billing/pending_bills.html', {
        'page_obj':      page_obj,
        'bills':         page_obj.object_list,
        'total_due':     total_due,
        'total_records': total_records,
        'customers':     customers,
        'q': q, 'customer': customer_f, 'date_from': date_from, 'date_to': date_to,
    })


# ─────────────────────────── settle payment ───────────────────────────────────

def settle_bill(request, pk):
    bill     = get_object_or_404(
        Bill.objects.select_related('customer').prefetch_related('payments'), pk=pk
    )
    customer = bill.customer

    if bill.status != Bill.PENDING:
        messages.info(request, f'Bill {bill.bill_number} is not pending — nothing to settle.')
        return redirect('billing:bill_detail', pk=pk)

    credit_balance = max(Decimal('0'), customer.balance)

    if request.method == 'POST':
        try:
            _settle_payment(request, bill, customer)
            bill.refresh_from_db()
            log_activity(
                request, 'billing', 'bill_settled',
                f"Payment recorded for bill: {bill.bill_number} | {bill.customer.name} | Status: {bill.get_status_display()}",
                reverse('billing:bill_detail', kwargs={'pk': bill.pk}),
            )
            if bill.status == Bill.COMPLETED:
                messages.success(request, f'Bill {bill.bill_number} fully settled and marked Completed.')
            else:
                messages.success(request, f'Partial payment recorded for {bill.bill_number}.')
            return redirect('billing:bill_detail', pk=pk)
        except ValueError as exc:
            messages.error(request, str(exc))

    return render(request, 'billing/settle_bill.html', {
        'bill':           bill,
        'customer':       customer,
        'credit_balance': credit_balance,
        'today':          timezone.localdate().isoformat(),
    })


@transaction.atomic
def _settle_payment(request, bill, customer):
    POST   = request.POST
    method = POST.get('settlement_method', '')

    valid  = {'FULL_CASH', 'FULL_CHEQUE', 'PARTIAL_CASH', 'PARTIAL_CHEQUE', 'MIXED'}
    if method not in valid:
        raise ValueError('Select a valid settlement method.')

    bill_obj     = Bill.objects.select_for_update().get(pk=bill.pk)
    customer_obj = Customer.objects.select_for_update().get(pk=customer.pk)

    amount_due   = bill_obj.amount_due
    credit       = max(Decimal('0'), customer_obj.balance)

    use_balance  = POST.get('use_credit_balance') == '1' and credit > 0
    balance_now  = min(credit, amount_due) if use_balance else Decimal('0')

    cash_in        = Decimal('0')
    cheque_in      = Decimal('0')
    cheque_records = []

    if method in ('FULL_CASH', 'PARTIAL_CASH', 'MIXED'):
        cash_in = _d(POST.get('cash_amount'))
    if method in ('FULL_CHEQUE', 'PARTIAL_CHEQUE', 'MIXED'):
        cheque_records, cheque_in = _parse_cheques(POST, customer_obj.name)

    collected  = cash_in + cheque_in
    effective  = (balance_now + collected).quantize(_TWO, rounding=ROUND_HALF_UP)

    if method in ('FULL_CASH', 'FULL_CHEQUE', 'MIXED'):
        if effective < amount_due:
            raise ValueError(
                f'Full settlement requires Rs. {amount_due:,.2f}. '
                f'You entered Rs. {effective:,.2f}.'
            )

    if collected <= 0 and balance_now <= 0:
        raise ValueError('Enter a payment amount or apply credit balance.')

    settle_date  = timezone.localdate()
    senovka      = POST.get('cash_is_senovka_transfer') == '1'
    pay_records  = []

    if balance_now > 0:
        pay_records.append({'method': Payment.BALANCE, 'amount': balance_now})
    if cash_in > 0:
        pay_records.append({'method': Payment.CASH, 'amount': cash_in, 'is_senovka_transfer': senovka})
    if cheque_records:
        pay_records.extend(cheque_records)

    for pd in pay_records:
        _save_payment(bill_obj, settle_date, pd)

    actual_applied = min(effective, amount_due)
    overpayment    = max(Decimal('0'), effective - amount_due)

    bill_obj.amount_paid  = (bill_obj.amount_paid  + collected ).quantize(_TWO, rounding=ROUND_HALF_UP)
    bill_obj.balance_used = (bill_obj.balance_used + balance_now).quantize(_TWO, rounding=ROUND_HALF_UP)
    bill_obj.amount_due   = max(Decimal('0'), (bill_obj.amount_due - effective).quantize(_TWO, rounding=ROUND_HALF_UP))

    if bill_obj.amount_due <= 0:
        bill_obj.status    = Bill.COMPLETED
        bill_obj.amount_due= Decimal('0')

    bill_obj.save(update_fields=['amount_paid', 'balance_used', 'amount_due', 'status'])

    # Cash/cheque received increases customer balance (reduces debt).
    # Credit balance used reduces it (spent on this bill).
    customer_obj.balance = customer_obj.balance + collected - balance_now
    customer_obj.save(update_fields=['balance'])


# ─────────────────────────── exports ──────────────────────────────────────────

def _filtered_export_qs(request):
    f  = _parse_filters(request)
    qs = _apply_filters(Bill.objects.select_related('customer').all(), f)
    return qs, f


def bill_export_excel(request):
    qs, _ = _filtered_export_qs(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bills'
    gen = timezone.localtime()
    total = qs.aggregate(s=Sum('total_amount'))['s'] or 0

    ws.append(['Billing Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Records', qs.count()])
    ws.append(['Grand Total (Rs.)', float(total)])
    ws.append([])

    headers = ['#', 'Bill No.', 'Customer', 'Date', 'Status', 'Method',
               'Subtotal', 'Discount', 'Total', 'Bal. Used', 'Paid', 'Due']
    ws.append(headers)
    hrow = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, b in enumerate(qs, 1):
        ws.append([
            i, b.bill_number, b.customer.name, b.bill_date.strftime('%d-%m-%Y'),
            b.get_status_display(), b.get_payment_method_display() if b.payment_method else '',
            float(b.subtotal), float(b.discount_amount), float(b.total_amount),
            float(b.balance_used), float(b.amount_paid), float(b.amount_due),
        ])

    for col, w in zip('ABCDEFGHIJKL', [4,16,22,12,12,18,12,10,12,10,10,10]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="bills_{ts}.xlsx"'
    wb.save(response)
    return response


def bill_export_pdf(request):
    qs, f = _filtered_export_qs(request)
    total = qs.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    due   = qs.aggregate(s=Sum('amount_due'))['s']   or Decimal('0')

    context = {
        'bills':          qs,
        'total_amount':   total,
        'total_due':      due,
        'total_records':  qs.count(),
        'generated_at':   timezone.localtime(),
        **f,
    }
    template = get_template('billing/bill_report_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="bills_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'billing/bill_report_pdf.html', context)


def bill_print(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related('customer', 'created_by')
            .prefetch_related('items__product', 'payments'),
        pk=pk,
    )
    customer_balance = bill.customer.balance
    net_value = bill.total_amount - customer_balance
    return render(request, 'billing/bill_print.html', {
        'bill':                 bill,
        'company':              CompanySettings.get(),
        'customer_balance':     customer_balance,
        'customer_balance_abs': abs(customer_balance),
        'net_value':            net_value,
    })


# ─────────────────────────── sales helpers ────────────────────────────────────

def _sales_filters(request):
    date_from, date_to = resolve_date_range(request)
    return {
        'q':         request.GET.get('q', '').strip(),
        'customer':  request.GET.get('customer', '').strip(),
        'date_from': date_from,
        'date_to':   date_to,
        'status':    request.GET.get('status', '').strip(),
    }


def _apply_sales_filters(qs, f):
    if f['q']:
        qs = qs.filter(Q(bill_number__icontains=f['q']) | Q(customer__name__icontains=f['q']))
    if f['customer']:
        qs = qs.filter(customer__name__icontains=f['customer'])
    if f['date_from']:
        qs = qs.filter(bill_date__gte=f['date_from'])
    if f['date_to']:
        qs = qs.filter(bill_date__lte=f['date_to'])
    if f['status']:
        qs = qs.filter(status=f['status'])
    return qs


def _build_filter_qs(request):
    """Return a querystring string without the 'page' key, for pagination links."""
    params = request.GET.copy()
    params.pop('page', None)
    return params.urlencode()


# ─────────────────────────── cash sales ───────────────────────────────────────

def _has_cash_sub():
    return Payment.objects.filter(bill=OuterRef('pk'), method=Payment.CASH)

def _has_cheque_sub():
    return Payment.objects.filter(bill=OuterRef('pk'), method=Payment.CHEQUE)

def _has_direct_cash_sub():
    return Payment.objects.filter(bill=OuterRef('pk'), method=Payment.CASH, is_senovka_transfer=False)

def _has_senovka_cash_sub():
    return Payment.objects.filter(bill=OuterRef('pk'), method=Payment.CASH, is_senovka_transfer=True)


def _cash_qs(request):
    f  = _sales_filters(request)
    qs = (
        Bill.objects.annotate(
            has_cash=Exists(_has_cash_sub()),
            has_cheque=Exists(_has_cheque_sub()),
            has_direct_cash=Exists(_has_direct_cash_sub()),
            has_senovka_cash=Exists(_has_senovka_cash_sub()),
        )
        .filter(has_cash=True)           # includes split bills
        .select_related('customer')
    )
    qs = _apply_sales_filters(qs, f).order_by('-bill_date', '-id')
    return qs, f


def cash_sales_list(request):
    qs, f      = _cash_qs(request)
    agg        = qs.aggregate(
        ta=Sum('total_amount'),
        tp=Sum('amount_paid'),
        td=Sum('amount_due'),
    )
    total_amount  = agg['ta'] or Decimal('0')
    total_paid    = agg['tp'] or Decimal('0')
    total_due     = agg['td'] or Decimal('0')
    total_records = qs.count()

    bill_ids = qs.values('id')
    direct_cash_total = Payment.objects.filter(
        bill_id__in=bill_ids, method=Payment.CASH, is_senovka_transfer=False,
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
    senovka_total = Payment.objects.filter(
        bill_id__in=bill_ids, method=Payment.CASH, is_senovka_transfer=True,
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0')

    paginator  = Paginator(qs, 20)
    page_obj   = paginator.get_page(request.GET.get('page'))
    customers  = Customer.objects.order_by('name')
    filter_qs  = _build_filter_qs(request)

    return render(request, 'billing/cash_sales.html', {
        'page_obj':           page_obj,
        'bills':              page_obj.object_list,
        'total_records':      total_records,
        'total_amount':       total_amount,
        'total_paid':         total_paid,
        'total_due':          total_due,
        'direct_cash_total':  direct_cash_total,
        'senovka_total':      senovka_total,
        'customers':          customers,
        'filter_qs':          filter_qs,
        **f,
    })


def cash_sales_export_excel(request):
    qs, _ = _cash_qs(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Cash Sales'
    gen = timezone.localtime()

    ws.append(['Cash Sales Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Records', qs.count()])
    ws.append(['Grand Total (Rs.)', float(qs.aggregate(s=Sum('total_amount'))['s'] or 0)])
    ws.append([])

    headers = ['#', 'Bill No.', 'Customer', 'Date', 'Status', 'Total (Rs.)', 'Paid (Rs.)', 'Due (Rs.)']
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, b in enumerate(qs, 1):
        ws.append([
            i,
            b.bill_number,
            b.customer.name,
            b.bill_date.strftime('%d-%m-%Y'),
            b.get_status_display(),
            float(b.total_amount),
            float(b.amount_paid),
            float(b.amount_due),
        ])

    for col, w in zip('ABCDEFGH', [4, 18, 24, 12, 12, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="cash_sales_{ts}.xlsx"'
    wb.save(response)
    return response


def cash_sales_export_pdf(request):
    qs, f = _cash_qs(request)
    agg   = qs.aggregate(
        ta=Sum('total_amount'),
        tp=Sum('amount_paid'),
        td=Sum('amount_due'),
    )
    context = {
        'bills':         qs,
        'total_amount':  agg['ta'] or Decimal('0'),
        'total_paid':    agg['tp'] or Decimal('0'),
        'total_due':     agg['td'] or Decimal('0'),
        'total_records': qs.count(),
        'generated_at':  timezone.localtime(),
        **f,
    }
    template = get_template('billing/cash_sales_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="cash_sales_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'billing/cash_sales_pdf.html', context)


# ─────────────────────────── cheque sales ─────────────────────────────────────

def _cheque_qs(request):
    f             = _sales_filters(request)
    cheque_status = request.GET.get('cheque_status', '').strip()
    qs = (
        Bill.objects.annotate(
            has_cash=Exists(_has_cash_sub()),
            has_cheque=Exists(_has_cheque_sub()),
        )
        .filter(has_cheque=True)         # includes split bills
        .select_related('customer')
        .prefetch_related('payments')
    )
    qs = _apply_sales_filters(qs, f)
    if cheque_status:
        qs = qs.filter(
            payments__method=Payment.CHEQUE,
            payments__cheque_status=cheque_status,
        ).distinct()
    qs = qs.order_by('-bill_date', '-id')
    return qs, f, cheque_status


def cheque_sales_list(request):
    qs, f, cheque_status = _cheque_qs(request)
    agg = qs.aggregate(
        ta=Sum('total_amount'),
        tp=Sum('amount_paid'),
        td=Sum('amount_due'),
    )
    total_amount  = agg['ta'] or Decimal('0')
    total_paid    = agg['tp'] or Decimal('0')
    total_due     = agg['td'] or Decimal('0')
    total_records = qs.count()

    today            = timezone.localdate()
    warning_date     = today + timedelta(days=3)
    pending_cheques  = Payment.objects.filter(method=Payment.CHEQUE, cheque_status=Payment.CHQ_PENDING)
    maturing_today_count = pending_cheques.filter(maturity_date=today).count()
    maturing_soon_count  = pending_cheques.filter(maturity_date__gt=today, maturity_date__lte=warning_date).count()

    paginator  = Paginator(qs, 20)
    page_obj   = paginator.get_page(request.GET.get('page'))
    customers  = Customer.objects.order_by('name')
    filter_qs  = _build_filter_qs(request)

    return render(request, 'billing/cheque_sales.html', {
        'page_obj':             page_obj,
        'bills':                page_obj.object_list,
        'total_records':        total_records,
        'total_amount':         total_amount,
        'total_paid':           total_paid,
        'total_due':            total_due,
        'customers':            customers,
        'cheque_status':        cheque_status,
        'filter_qs':            filter_qs,
        'today':                today,
        'warning_date':         warning_date,
        'maturing_today_count': maturing_today_count,
        'maturing_soon_count':  maturing_soon_count,
        **f,
    })


def cheque_sales_export_excel(request):
    qs, _, cheque_status = _cheque_qs(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Cheque Sales'
    gen = timezone.localtime()

    ws.append(['Cheque Sales Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Records', qs.count()])
    ws.append([])

    headers = [
        '#', 'Bill No.', 'Customer', 'Date', 'Status',
        'Total (Rs.)', 'Paid (Rs.)', 'Due (Rs.)',
        'Cheque No.', 'Bank', 'Branch', 'Maturity Date', 'Cheque Status',
    ]
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, b in enumerate(qs, 1):
        chq = b.payments.filter(method=Payment.CHEQUE).first()
        ws.append([
            i,
            b.bill_number,
            b.customer.name,
            b.bill_date.strftime('%d-%m-%Y'),
            b.get_status_display(),
            float(b.total_amount),
            float(b.amount_paid),
            float(b.amount_due),
            chq.cheque_number  if chq else '',
            chq.bank_name      if chq else '',
            chq.branch_name    if chq else '',
            chq.maturity_date.strftime('%d-%m-%Y') if chq and chq.maturity_date else '',
            chq.cheque_status  if chq else '',
        ])

    for col, w in zip('ABCDEFGHIJKLM', [4, 18, 22, 12, 12, 13, 13, 13, 16, 20, 18, 14, 14]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="cheque_sales_{ts}.xlsx"'
    wb.save(response)
    return response


def cheque_sales_export_pdf(request):
    qs, f, cheque_status = _cheque_qs(request)
    agg = qs.aggregate(
        ta=Sum('total_amount'),
        tp=Sum('amount_paid'),
        td=Sum('amount_due'),
    )
    context = {
        'bills':         qs,
        'total_amount':  agg['ta'] or Decimal('0'),
        'total_paid':    agg['tp'] or Decimal('0'),
        'total_due':     agg['td'] or Decimal('0'),
        'total_records': qs.count(),
        'generated_at':  timezone.localtime(),
        'cheque_status': cheque_status,
        **f,
    }
    template = get_template('billing/cheque_sales_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="cheque_sales_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'billing/cheque_sales_pdf.html', context)


@transaction.atomic
def cheque_update_status(request, pk):
    if request.method != 'POST':
        return redirect(reverse('billing:cheque_sales'))

    valid_statuses = {Payment.CHQ_PENDING, Payment.CHQ_CLEARED, Payment.CHQ_BOUNCED, Payment.CHQ_HOLD}
    new_status = request.POST.get('status', '').strip()
    if new_status not in valid_statuses:
        messages.error(request, 'Invalid cheque status.')
        return redirect(reverse('billing:cheque_sales'))

    payment    = get_object_or_404(Payment.objects.select_for_update(), pk=pk, method=Payment.CHEQUE)
    old_status = payment.cheque_status

    # ── Bounce: reverse the financial impact ─────────────────────────────────
    if new_status == Payment.CHQ_BOUNCED and old_status != Payment.CHQ_BOUNCED:
        chq_amount = payment.amount

        if payment.bill_id:
            bill     = Bill.objects.select_for_update().get(pk=payment.bill_id)
            customer = Customer.objects.select_for_update().get(pk=bill.customer_id)

            bill.amount_paid = max(Decimal('0'), bill.amount_paid - chq_amount)
            max_possible_due = bill.total_amount - bill.balance_used
            new_due          = min(max_possible_due, bill.amount_due + chq_amount)
            overpayment_to_reverse = (bill.amount_due + chq_amount) - new_due
            bill.amount_due = new_due

            if bill.amount_due > 0 and bill.status == Bill.COMPLETED:
                bill.status = Bill.PENDING
            bill.save(update_fields=['amount_paid', 'amount_due', 'status'])

            if overpayment_to_reverse > 0:
                customer.balance = customer.balance - overpayment_to_reverse
                customer.save(update_fields=['balance'])
        else:
            customer = Customer.objects.select_for_update().get(pk=payment.customer_id)
            customer.balance -= chq_amount
            customer.save(update_fields=['balance'])
            add_ledger_entry(
                customer,
                date=timezone.localdate(),
                description=f'Cheque Bounced — Direct Settlement (Cheque #{payment.cheque_number or payment.pk})',
                transaction_type=CustomerLedger.CHEQUE_BOUNCED,
                debit=chq_amount,
            )

    # ── Re-clear a previously bounced cheque: re-apply the payment ───────────
    elif new_status == Payment.CHQ_CLEARED and old_status == Payment.CHQ_BOUNCED:
        chq_amount = payment.amount

        if payment.bill_id:
            bill     = Bill.objects.select_for_update().get(pk=payment.bill_id)
            customer = Customer.objects.select_for_update().get(pk=bill.customer_id)

            old_due          = bill.amount_due
            bill.amount_paid = bill.amount_paid + chq_amount
            bill.amount_due  = max(Decimal('0'), bill.amount_due - chq_amount)
            overpayment      = max(Decimal('0'), chq_amount - old_due)

            if bill.amount_due <= 0:
                bill.status     = Bill.COMPLETED
                bill.amount_due = Decimal('0')
            bill.save(update_fields=['amount_paid', 'amount_due', 'status'])

            if overpayment > 0:
                customer.balance = customer.balance + overpayment
                customer.save(update_fields=['balance'])
        else:
            customer = Customer.objects.select_for_update().get(pk=payment.customer_id)
            customer.balance += chq_amount
            customer.save(update_fields=['balance'])
            add_ledger_entry(
                customer,
                date=timezone.localdate(),
                description=f'Cheque Re-Cleared — Direct Settlement (Cheque #{payment.cheque_number or payment.pk})',
                transaction_type=CustomerLedger.CHEQUE_CLEARED,
                credit=chq_amount,
            )

    payment.cheque_status = new_status
    payment.save(update_fields=['cheque_status'])

    label = {
        Payment.CHQ_PENDING: 'Pending',
        Payment.CHQ_CLEARED: 'Deposited / Cleared',
        Payment.CHQ_BOUNCED: 'Bounced',
        Payment.CHQ_HOLD:    'On Hold',
    }
    chq_ref = payment.cheque_number or f'#{payment.pk}'
    messages.success(request, f"Cheque {chq_ref} marked as {label[new_status]}.")

    next_url = request.POST.get('next', '').strip()
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect(reverse('billing:cheque_sales'))


def _redirect_after_payment_action(request, payment):
    next_url = request.POST.get('next', '').strip()
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    if payment.bill_id:
        return redirect('billing:bill_detail', pk=payment.bill_id)
    return redirect('billing:cheque_sales')


def payment_update_dates(request, pk):
    """Edit only a cheque's received/maturity date — no financial impact."""
    payment = get_object_or_404(Payment, pk=pk, method=Payment.CHEQUE)
    if request.method != 'POST':
        return _redirect_after_payment_action(request, payment)

    from datetime import datetime
    def _parse(val):
        try:
            return datetime.strptime(val, '%Y-%m-%d').date() if val else None
        except ValueError:
            return None

    maturity_date = _parse(request.POST.get('maturity_date', ''))
    if not maturity_date:
        messages.error(request, 'Maturity date is required.')
        return _redirect_after_payment_action(request, payment)

    payment.received_date = _parse(request.POST.get('received_date', ''))
    payment.maturity_date = maturity_date
    payment.save(update_fields=['received_date', 'maturity_date'])

    chq_ref = payment.cheque_number or f'#{payment.pk}'
    log_activity(
        request, 'billing', 'cheque_dates_updated',
        f"Cheque {chq_ref} dates updated: received {payment.received_date}, matures {payment.maturity_date}",
        reverse('billing:bill_detail', kwargs={'pk': payment.bill_id}) if payment.bill_id else reverse('billing:cheque_sales'),
    )
    messages.success(request, f'Cheque {chq_ref} dates updated.')
    return _redirect_after_payment_action(request, payment)


@superadmin_required
@transaction.atomic
def payment_delete(request, pk):
    """Delete a cheque payment and reverse its financial impact on the bill /
    customer balance. If the cheque was already marked BOUNCED, its amount was
    already reversed at that point, so deleting it just removes the row."""
    payment = get_object_or_404(Payment.objects.select_for_update(), pk=pk, method=Payment.CHEQUE)
    if request.method != 'POST':
        return _redirect_after_payment_action(request, payment)

    chq_amount   = payment.amount
    chq_ref      = payment.cheque_number or f'#{payment.pk}'
    was_reversed = payment.cheque_status == Payment.CHQ_BOUNCED
    bill_id      = payment.bill_id

    if not was_reversed:
        if payment.bill_id:
            bill     = Bill.objects.select_for_update().get(pk=payment.bill_id)
            customer = Customer.objects.select_for_update().get(pk=bill.customer_id)

            bill.amount_paid = max(Decimal('0'), bill.amount_paid - chq_amount)
            max_possible_due = bill.total_amount - bill.balance_used
            new_due          = min(max_possible_due, bill.amount_due + chq_amount)
            overpayment_to_reverse = (bill.amount_due + chq_amount) - new_due
            bill.amount_due = new_due

            if bill.amount_due > 0 and bill.status == Bill.COMPLETED:
                bill.status = Bill.PENDING
            bill.save(update_fields=['amount_paid', 'amount_due', 'status'])

            if overpayment_to_reverse > 0:
                customer.balance = customer.balance - overpayment_to_reverse
                customer.save(update_fields=['balance'])
        else:
            customer = Customer.objects.select_for_update().get(pk=payment.customer_id)
            customer.balance -= chq_amount
            customer.save(update_fields=['balance'])
            add_ledger_entry(
                customer,
                date=timezone.localdate(),
                description=f'Cheque Deleted — Direct Settlement (Cheque #{chq_ref})',
                transaction_type=CustomerLedger.MANUAL_ADJUSTMENT,
                debit=chq_amount,
            )

    payment.delete()
    log_activity(
        request, 'billing', 'cheque_deleted',
        f"Cheque {chq_ref} deleted"
        + ('' if was_reversed else f", Rs. {chq_amount} reversed"),
        reverse('billing:bill_detail', kwargs={'pk': bill_id}) if bill_id else reverse('billing:cheque_sales'),
    )
    messages.success(
        request,
        f'Cheque {chq_ref} deleted.' if was_reversed
        else f'Cheque {chq_ref} deleted and Rs. {chq_amount:,.2f} reversed.',
    )

    if bill_id:
        return redirect('billing:bill_detail', pk=bill_id)
    return redirect('billing:cheque_sales')


# ─────────────────────────── split sales ──────────────────────────────────────

def _split_qs(request):
    f  = _sales_filters(request)
    qs = (
        Bill.objects.annotate(
            has_cash=Exists(_has_cash_sub()),
            has_cheque=Exists(_has_cheque_sub()),
        )
        .filter(has_cash=True, has_cheque=True)
        .select_related('customer')
        .prefetch_related('payments')
    )
    qs = _apply_sales_filters(qs, f).order_by('-bill_date', '-id')
    return qs, f


def split_sales_list(request):
    qs, f = _split_qs(request)
    agg   = qs.aggregate(
        ta=Sum('total_amount'),
        tp=Sum('amount_paid'),
        td=Sum('amount_due'),
        tb=Sum('balance_used'),
    )
    total_amount  = agg['ta'] or Decimal('0')
    total_paid    = agg['tp'] or Decimal('0')
    total_due     = agg['td'] or Decimal('0')
    total_bal     = agg['tb'] or Decimal('0')
    total_records = qs.count()

    paginator  = Paginator(qs, 20)
    page_obj   = paginator.get_page(request.GET.get('page'))
    customers  = Customer.objects.order_by('name')
    filter_qs  = _build_filter_qs(request)

    return render(request, 'billing/split_sales.html', {
        'page_obj':      page_obj,
        'bills':         page_obj.object_list,
        'total_records': total_records,
        'total_amount':  total_amount,
        'total_paid':    total_paid,
        'total_due':     total_due,
        'total_bal':     total_bal,
        'customers':     customers,
        'filter_qs':     filter_qs,
        **f,
    })


def split_sales_export_excel(request):
    qs, _ = _split_qs(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl required.', status=500)

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Split Sales'
    gen = timezone.localtime()

    ws.append(['Split Sales Report — SENOVKA ERP'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Records', qs.count()])
    ws.append([])

    headers = [
        '#', 'Bill No.', 'Customer', 'Date', 'Status',
        'Total (Rs.)', 'Cash Paid (Rs.)', 'Cheque Paid (Rs.)',
        'Balance Used (Rs.)', 'Total Paid (Rs.)', 'Due (Rs.)',
    ]
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, b in enumerate(qs, 1):
        cash_paid   = sum(
            p.amount for p in b.payments.all() if p.method == Payment.CASH
        )
        cheque_paid = sum(
            p.amount for p in b.payments.all() if p.method == Payment.CHEQUE
        )
        ws.append([
            i,
            b.bill_number,
            b.customer.name,
            b.bill_date.strftime('%d-%m-%Y'),
            b.get_status_display(),
            float(b.total_amount),
            float(cash_paid),
            float(cheque_paid),
            float(b.balance_used),
            float(b.amount_paid),
            float(b.amount_due),
        ])

    for col, w in zip('ABCDEFGHIJK', [4, 18, 22, 12, 12, 13, 15, 16, 16, 14, 12]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="split_sales_{ts}.xlsx"'
    wb.save(response)
    return response


def split_sales_export_pdf(request):
    qs, f = _split_qs(request)
    agg   = qs.aggregate(
        ta=Sum('total_amount'),
        tp=Sum('amount_paid'),
        td=Sum('amount_due'),
        tb=Sum('balance_used'),
    )
    context = {
        'bills':         qs,
        'total_amount':  agg['ta'] or Decimal('0'),
        'total_paid':    agg['tp'] or Decimal('0'),
        'total_due':     agg['td'] or Decimal('0'),
        'total_bal':     agg['tb'] or Decimal('0'),
        'total_records': qs.count(),
        'generated_at':  timezone.localtime(),
        **f,
    }
    template = get_template('billing/split_sales_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="split_sales_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'billing/split_sales_pdf.html', context)
