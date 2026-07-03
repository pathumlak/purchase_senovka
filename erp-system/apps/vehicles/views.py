from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone

from .models import VehicleLog

_ONE = Decimal('0.1')


def _d1(val):
    try:
        return Decimal(str(val or '0')).quantize(_ONE, rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError):
        return Decimal('0')


def vehicle_log_list(request):
    today = timezone.localdate()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12) or year < 2000:
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month

    if request.method == 'POST':
        return _handle_create(request, year, month)

    qs          = VehicleLog.objects.filter(date__year=year, date__month=month)
    total_km    = qs.aggregate(s=Sum('total_km'))['s'] or Decimal('0')
    total_trips = qs.count()

    # Month navigation
    prev_month = month - 1 or 12
    prev_year  = year - 1 if month == 1 else year
    next_month = month % 12 + 1
    next_year  = year + 1 if month == 12 else year

    month_name = date(year, month, 1).strftime('%B')

    return render(request, 'vehicles/vehicle_log.html', {
        'logs':        qs,
        'year':        year,
        'month':       month,
        'month_name':  month_name,
        'total_km':    total_km,
        'total_trips': total_trips,
        'prev_year':   prev_year,
        'prev_month':  prev_month,
        'next_year':   next_year,
        'next_month':  next_month,
        'today':       today.isoformat(),
        'drivers':     VehicleLog.objects.values_list('driver_name', flat=True).distinct().order_by('driver_name'),
    })


def _handle_create(request, year, month):
    POST = request.POST
    errors = []

    driver   = POST.get('driver_name', '').strip()
    from_loc = POST.get('from_location', '').strip()
    to_loc   = POST.get('to_location', '').strip()
    purpose  = POST.get('purpose', '').strip()
    date_str = POST.get('date', '').strip()
    start_km = _d1(POST.get('start_km'))
    end_km   = _d1(POST.get('end_km'))

    if not driver:   errors.append('Driver name is required.')
    if not from_loc: errors.append('From location is required.')
    if not to_loc:   errors.append('To location is required.')
    if not purpose:  errors.append('Purpose is required.')
    if end_km <= start_km:
        errors.append('End KM must be greater than Start KM.')

    try:
        from datetime import datetime
        log_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        log_date = timezone.localdate()

    if errors:
        for e in errors:
            messages.error(request, e)
    else:
        VehicleLog.objects.create(
            date          = log_date,
            driver_name   = driver,
            from_location = from_loc,
            to_location   = to_loc,
            start_km      = start_km,
            end_km        = end_km,
            purpose       = purpose,
            created_by    = request.user if request.user.is_authenticated else None,
        )
        messages.success(request, 'Trip logged successfully.')

    return redirect(f'/vehicles/?year={log_date.year}&month={log_date.month}')


def vehicle_log_edit(request, pk):
    log = get_object_or_404(VehicleLog, pk=pk)
    if request.method != 'POST':
        return redirect('vehicles:list')

    POST = request.POST
    from datetime import datetime

    driver   = POST.get('driver_name', '').strip()
    from_loc = POST.get('from_location', '').strip()
    to_loc   = POST.get('to_location', '').strip()
    purpose  = POST.get('purpose', '').strip()
    date_str = POST.get('date', '').strip()
    start_km = _d1(POST.get('start_km'))
    end_km   = _d1(POST.get('end_km'))

    errors = []
    if not driver:   errors.append('Driver name is required.')
    if not from_loc: errors.append('From location is required.')
    if not to_loc:   errors.append('To location is required.')
    if not purpose:  errors.append('Purpose is required.')
    if end_km <= start_km:
        errors.append('End KM must be greater than Start KM.')

    try:
        log_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        log_date = log.date

    if errors:
        for e in errors:
            messages.error(request, e)
    else:
        log.date          = log_date
        log.driver_name   = driver
        log.from_location = from_loc
        log.to_location   = to_loc
        log.start_km      = start_km
        log.end_km        = end_km
        log.purpose       = purpose
        log.save()
        messages.success(request, 'Trip updated.')

    year  = POST.get('year',  log_date.year)
    month = POST.get('month', log_date.month)
    return redirect(f'/vehicles/?year={year}&month={month}')


def vehicle_log_delete(request, pk):
    log = get_object_or_404(VehicleLog, pk=pk)
    if request.method == 'POST':
        year  = request.POST.get('year',  log.date.year)
        month = request.POST.get('month', log.date.month)
        log.delete()
        messages.success(request, 'Trip deleted.')
        return redirect(f'/vehicles/?year={year}&month={month}')
    return redirect('vehicles:list')


# ─────────────────────────── exports ──────────────────────────────────────────

def _export_qs(request):
    today = timezone.localdate()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12) or year < 2000:
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month

    qs         = VehicleLog.objects.filter(date__year=year, date__month=month)
    total_km   = qs.aggregate(s=Sum('total_km'))['s'] or Decimal('0')
    month_name = date(year, month, 1).strftime('%B')
    return qs, total_km, month_name, year, month


def vehicle_export_excel(request):
    qs, total_km, month_name, year, month = _export_qs(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl is required for Excel export.', status=500)

    wb  = Workbook()
    ws  = wb.active
    ws.title = f'Vehicle Log {month_name} {year}'
    gen = timezone.localtime()

    # Title block
    ws.append([f'Vehicle Maintenance — {month_name} {year}'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(['Generated At', gen.strftime('%d %b %Y %I:%M %p')])
    ws.append(['Total Trips', qs.count()])
    ws.append(['Total KM', float(total_km)])
    ws.append([])

    # Header row
    headers = ['#', 'Date', 'Driver', 'From', 'To', 'Start KM', 'End KM', 'Total KM', 'Purpose']
    ws.append(headers)
    hrow  = ws.max_row
    hfill = PatternFill('solid', fgColor='1D4ED8')
    hfont = Font(bold=True, color='FFFFFF')
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=hrow, column=ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, log in enumerate(qs, 1):
        ws.append([
            i,
            log.date.strftime('%d-%m-%Y'),
            log.driver_name,
            log.from_location,
            log.to_location,
            float(log.start_km),
            float(log.end_km),
            float(log.total_km),
            log.purpose,
        ])
        for ci in range(1, 10):
            ws.cell(row=ws.max_row, column=ci).border = border

    # Total row
    ws.append(['', '', '', '', 'TOTAL', '', '', float(total_km), ''])
    trow = ws.max_row
    for ci in range(1, 10):
        cell = ws.cell(row=trow, column=ci)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='F0FDF4')
    ws.cell(row=trow, column=8).font = Font(bold=True, color='15803D')

    # Column widths
    for col, width in zip('ABCDEFGHI', [4, 14, 20, 22, 22, 11, 11, 11, 30]):
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = gen.strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="vehicle_log_{year}_{month:02d}_{ts}.xlsx"'
    wb.save(response)
    return response


def vehicle_export_pdf(request):
    qs, total_km, month_name, year, month = _export_qs(request)

    context = {
        'logs':         qs,
        'total_km':     total_km,
        'total_trips':  qs.count(),
        'month_name':   month_name,
        'year':         year,
        'generated_at': timezone.localtime(),
    }
    template = get_template('vehicles/vehicle_log_pdf.html')
    html     = template.render(context)

    try:
        from xhtml2pdf import pisa
        response = HttpResponse(content_type='application/pdf')
        ts = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="vehicle_log_{year}_{month:02d}_{ts}.pdf"'
        status = pisa.CreatePDF(html, dest=response)
        if status.err:
            return HttpResponse('PDF generation error.', status=500)
        return response
    except ImportError:
        context['printable'] = True
        return render(request, 'vehicles/vehicle_log_pdf.html', context)
